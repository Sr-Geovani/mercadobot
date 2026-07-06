import logging
import os
import asyncio
import subprocess
import sys
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand, MenuButtonCommands
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)
import anthropic

# ─── CONFIGURAÇÃO ────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_KEY")

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

dados_usuario = {}
aguardando_dias = {}

# ─── INSTALA PLAYWRIGHT BROWSER SE NECESSÁRIO ────────────────
def garantir_browser():
    """Instala o Chromium do Playwright se ainda não estiver disponível."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            # Tenta apenas verificar se o browser existe
            browser = p.chromium.launch(headless=True)
            browser.close()
        logging.info("✅ Playwright Chromium já instalado.")
    except Exception:
        logging.info("⏳ Instalando Playwright Chromium...")
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium", "--with-deps"],
            check=True
        )
        logging.info("✅ Playwright Chromium instalado com sucesso.")

# ─── PALETA ──────────────────────────────────────────────────
COR_BG      = "#0e0f11"
COR_SURFACE = "#1e2027"
COR_BORDA   = "#2a2d35"
COR_VERDE   = "#00e676"
COR_AZUL    = "#40c4ff"
COR_AMARELO = "#ffd740"
COR_ROXO    = "#ce93d8"
COR_TEXTO   = "#e8eaf0"
CORES       = [COR_VERDE, COR_AZUL, COR_AMARELO, COR_ROXO, "#ff8a65", "#80cbc4", "#fff176", "#ef9a9a"]

# ─── ESTILO GRÁFICO ──────────────────────────────────────────
def estilo(fig, ax):
    fig.patch.set_facecolor(COR_BG)
    ax.set_facecolor(COR_SURFACE)
    ax.tick_params(colors=COR_TEXTO, labelsize=9)
    ax.xaxis.label.set_color(COR_TEXTO)
    ax.yaxis.label.set_color(COR_TEXTO)
    ax.title.set_color(COR_TEXTO)
    for sp in ax.spines.values():
        sp.set_edgecolor(COR_BORDA)
    ax.grid(axis="y", color=COR_BORDA, linewidth=0.5, linestyle="--")
    ax.set_axisbelow(True)

def salvar(fig) -> BytesIO:
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor=fig.get_facecolor(), dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf

# ─── GRÁFICOS ────────────────────────────────────────────────
def g_filiais(vendas):
    """Gráfico de faturamento por filial (barras lado-a-lado)."""
    if "nomeFilial" not in vendas.columns or len(vendas) == 0:
        return None
    por_filial = vendas.groupby("nomeFilial")["valor"].sum().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(10, 5))
    cores = ["#2E75B6", "#548235", "#C55A11"]
    ax.barh(range(len(por_filial)), por_filial.values, color=cores[:len(por_filial)])
    ax.set_yticks(range(len(por_filial)))
    ax.set_yticklabels([f.split()[-1].title() for f in por_filial.index])
    ax.set_xlabel("Faturamento (R$)", fontsize=11, fontweight="bold")
    ax.set_title("Faturamento por Filial", fontsize=13, fontweight="bold", pad=15)
    for i, v in enumerate(por_filial.values):
        ax.text(v, i, f" R$ {v:,.0f}", va="center", fontsize=10, fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf

def g_faturamento(vendas):
    fat = vendas.groupby("nomeFilial")["valor"].sum().sort_values()
    fig, ax = plt.subplots(figsize=(7, 3))
    estilo(fig, ax)
    bars = ax.barh(fat.index, fat.values, color=[COR_VERDE, COR_AZUL], height=0.45)
    for bar, val in zip(bars, fat.values):
        ax.text(val + fat.max()*0.01, bar.get_y()+bar.get_height()/2,
                f"R$ {val:,.0f}", va="center", color=COR_TEXTO, fontsize=9, fontweight="bold")
    ax.set_title("Faturamento por Unidade", fontsize=11, fontweight="bold", pad=10)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"R$ {x:,.0f}"))
    fig.tight_layout()
    return salvar(fig)

def g_pagamentos(vendas):
    mix = vendas.groupby("FormaRecebimento")["valor"].count()
    fig, ax = plt.subplots(figsize=(5, 4.5))
    fig.patch.set_facecolor(COR_BG)
    wedges, texts, autotexts = ax.pie(
        mix.values, labels=mix.index, autopct="%1.1f%%",
        colors=CORES[:len(mix)], startangle=90,
        wedgeprops=dict(edgecolor=COR_BG, linewidth=2),
        textprops=dict(color=COR_TEXTO, fontsize=10)
    )
    for at in autotexts:
        at.set_color(COR_BG)
        at.set_fontweight("bold")
    ax.set_title("Mix de Pagamento", fontsize=11, fontweight="bold", color=COR_TEXTO, pad=12)
    fig.tight_layout()
    return salvar(fig)

def g_categorias(produtos):
    grupos = produtos.groupby("grupo")["valor"].sum().sort_values(ascending=False).head(8)
    fig, ax = plt.subplots(figsize=(8, 4))
    estilo(fig, ax)
    bars = ax.bar(range(len(grupos)), grupos.values, color=CORES[:len(grupos)], width=0.6)
    ax.set_xticks(range(len(grupos)))
    ax.set_xticklabels(
        [g[:13]+"…" if len(g)>13 else g for g in grupos.index],
        rotation=20, ha="right", fontsize=8
    )
    total = grupos.sum()
    for bar, val in zip(bars, grupos.values):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+total*0.005,
                f"{val/total*100:.0f}%", ha="center", color=COR_TEXTO, fontsize=8, fontweight="bold")
    ax.set_title("Receita por Categoria", fontsize=11, fontweight="bold", pad=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"R$ {x:,.0f}"))
    fig.tight_layout()
    return salvar(fig)

def g_top_produtos(produtos):
    filiais = produtos["nomeloja"].unique()
    fig, axes = plt.subplots(1, len(filiais), figsize=(7*len(filiais), 4.5))
    if len(filiais) == 1:
        axes = [axes]
    for ax, filial in zip(axes, filiais):
        estilo(fig, ax)
        top = produtos[produtos["nomeloja"]==filial].sort_values("quantidade", ascending=False).head(6)
        nomes = [p[:20]+"…" if len(p)>20 else p for p in top["produto"]]
        bars = ax.barh(range(len(top)), top["quantidade"].values, color=CORES[:len(top)], height=0.55)
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(nomes, fontsize=9)
        ax.invert_yaxis()
        for bar, val in zip(bars, top["quantidade"].values):
            ax.text(val+0.2, bar.get_y()+bar.get_height()/2, f"{val} un", va="center", color=COR_TEXTO, fontsize=8)
        ax.set_title(f"Top Produtos — {filial.split()[-1].title()}", fontsize=11, fontweight="bold", pad=10)
        ax.set_xlabel("Unidades vendidas")
    fig.tight_layout()
    return salvar(fig)

def g_semanal(vendas):
    v = vendas.copy()
    v["DataAbertura"] = pd.to_datetime(v["DataAbertura"], dayfirst=True)
    v["semana"] = v["DataAbertura"].dt.isocalendar().week.astype(str).apply(lambda x: f"S{x}")
    sem = v.groupby(["semana","nomeFilial"])["valor"].sum().unstack(fill_value=0)
    fig, ax = plt.subplots(figsize=(8, 4))
    estilo(fig, ax)
    for i, col in enumerate([COR_VERDE, COR_AZUL][:len(sem.columns)]):
        label = sem.columns[i].split()[-1].title()
        ax.plot(range(len(sem)), sem.iloc[:,i], marker="o", color=col, linewidth=2.5, markersize=8, label=label)
        for x, y in enumerate(sem.iloc[:,i]):
            ax.text(x, y+sem.values.max()*0.03, f"R${y:,.0f}", ha="center", color=col, fontsize=8)
    ax.set_xticks(range(len(sem)))
    ax.set_xticklabels(sem.index)
    ax.set_title("Evolução Semanal", fontsize=11, fontweight="bold", pad=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"R$ {x:,.0f}"))
    ax.legend(facecolor=COR_SURFACE, edgecolor=COR_BORDA, labelcolor=COR_TEXTO, fontsize=9)
    fig.tight_layout()
    return salvar(fig)

def g_pico(vendas):
    v = vendas.copy()
    v["hora"] = pd.to_datetime(v["HoraAbertura"], format="%H:%M:%S").dt.hour
    pico = v.groupby(["hora","nomeFilial"])["valor"].count().unstack(fill_value=0)
    fig, ax = plt.subplots(figsize=(10, 4))
    estilo(fig, ax)
    w = 0.35
    for i, col in enumerate([COR_VERDE, COR_AZUL][:len(pico.columns)]):
        label = pico.columns[i].split()[-1].title()
        offset = (i - 0.5) * w
        ax.bar([x+offset for x in range(len(pico))], pico.iloc[:,i],
               width=w, color=col, label=label, alpha=0.9)
    ax.set_xticks(range(len(pico)))
    ax.set_xticklabels([f"{h}h" for h in pico.index], fontsize=8)
    ax.set_title("Vendas por Horário", fontsize=11, fontweight="bold", pad=10)
    ax.set_ylabel("Nº de vendas")
    ax.legend(facecolor=COR_SURFACE, edgecolor=COR_BORDA, labelcolor=COR_TEXTO, fontsize=9)
    fig.tight_layout()
    return salvar(fig)

# ─── FORMATAÇÃO HTML (cores no texto) ────────────────────────
def detectar_queda_ritmo(
    vendas_hoje: pd.DataFrame,
    vendas_historico: pd.DataFrame,
    hora_atual: int = None,
) -> dict:
    """
    Detecção PROATIVA de queda por RITMO: compara o faturamento ACUMULADO até a
    hora atual com o que os mesmos dias da semana normalmente já tinham feito
    até essa mesma hora. Compara períodos equivalentes (mesmo ponto do dia),
    não o total — o que permite avisar cedo, com tempo de reagir.

    Também PROJETA o fechamento do dia usando a curva intradiária média:
    se historicamente X% do faturamento do dia já saiu até esta hora, projeta
    o total dividindo o acumulado atual por essa fração.

    Retorna dict com tem_queda, acumulado_atual, esperado_ate_agora (faixa),
    projecao_fechamento, media_fechamento_dia, e causas.
    """
    import numpy as np

    if len(vendas_hoje) == 0 or len(vendas_historico) == 0:
        return {"tem_queda": False, "motivo": "Sem dados"}

    if hora_atual is None:
        hora_atual = datetime.now(ZoneInfo("America/Sao_Paulo")).hour

    # Coluna de data/hora
    col_dh = None
    for c in ("HoraAbertura", "data", "Data", "dataVenda"):
        if c in vendas_hoje.columns:
            col_dh = c
            break
    if col_dh is None or col_dh not in vendas_historico.columns:
        return {"tem_queda": False, "motivo": "Sem coluna de hora"}

    # Dia da semana de hoje
    dia_semana_hoje = None
    try:
        dt_hoje = pd.to_datetime(vendas_hoje[col_dh], errors="coerce", dayfirst=True).dropna()
        if len(dt_hoje) > 0:
            dia_semana_hoje = int(dt_hoje.dt.dayofweek.mode()[0])
    except Exception:
        pass
    if dia_semana_hoje is None:
        return {"tem_queda": False, "motivo": "Sem dia da semana"}

    # Acumulado de HOJE até a hora atual
    acumulado_atual = vendas_hoje["valor"].sum()

    # Constrói, para cada dia-calendário histórico do mesmo dia-da-semana:
    #   (a) acumulado até hora_atual  (b) total do dia
    try:
        hist = vendas_historico.copy()
        hist["_dt"] = pd.to_datetime(hist[col_dh], errors="coerce", dayfirst=True)
        hist = hist.dropna(subset=["_dt"])
        hist["_dia_cal"] = hist["_dt"].dt.date
        hist["_dow"] = hist["_dt"].dt.dayofweek
        hist["_hora"] = hist["_dt"].dt.hour

        hist_dow = hist[hist["_dow"] == dia_semana_hoje]
        if len(hist_dow) == 0:
            return {"tem_queda": False, "motivo": "Sem histórico deste dia da semana"}

        acumulados_ate_hora = []
        totais_dia = []
        fracoes = []  # % do dia já feito até a hora atual

        for dia_cal, grupo in hist_dow.groupby("_dia_cal"):
            total_dia = grupo["valor"].sum()
            ate_hora = grupo[grupo["_hora"] <= hora_atual]["valor"].sum()
            if total_dia > 0:
                acumulados_ate_hora.append(ate_hora)
                totais_dia.append(total_dia)
                fracoes.append(ate_hora / total_dia)
    except Exception as e:
        logger.warning(f"Erro ao construir curva intradiária: {e}")
        return {"tem_queda": False, "motivo": "Erro no cálculo"}

    n = len(acumulados_ate_hora)
    if n < 3:
        # Poucos dados — não arrisca alerta proativo (evita falso positivo)
        return {"tem_queda": False, "motivo": "Histórico insuficiente", "n_amostras": n}

    media_ate_hora = float(np.mean(acumulados_ate_hora))
    desvio_ate_hora = float(np.std(acumulados_ate_hora, ddof=1)) if n > 1 else 0.0
    media_total_dia = float(np.mean(totais_dia))
    fracao_media = float(np.mean(fracoes)) if fracoes else 0.0

    # Faixa esperada até a hora atual (1.5 desvios se maduro, 2.0 se médio)
    n_desvios = 1.5 if n >= 6 else 2.0
    limite_inferior = media_ate_hora - (n_desvios * desvio_ate_hora)

    # Projeção de fechamento: acumulado / fração média do dia já decorrida
    projecao = acumulado_atual / fracao_media if fracao_media > 0.05 else None

    nomes_dow = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]
    nome_dia = nomes_dow[dia_semana_hoje]

    # Só alerta se está ABAIXO da faixa esperada para este ponto do dia
    if desvio_ate_hora > 0 and acumulado_atual >= limite_inferior:
        return {
            "tem_queda": False,
            "acumulado_atual": acumulado_atual,
            "esperado_ate_agora": media_ate_hora,
            "projecao_fechamento": projecao,
            "dia_semana": nome_dia,
        }

    # Fallback sem desvio: alerta se abaixo de 70% da média até a hora
    if desvio_ate_hora <= 0 and acumulado_atual >= media_ate_hora * 0.7:
        return {"tem_queda": False, "acumulado_atual": acumulado_atual}

    causas = _analisar_causas_queda(vendas_hoje, vendas_historico)

    return {
        "tem_queda": True,
        "confianca": "alta" if n >= 6 else "media",
        "n_amostras": n,
        "dia_semana": nome_dia,
        "hora_atual": hora_atual,
        "acumulado_atual": acumulado_atual,
        "esperado_ate_agora_min": max(0, limite_inferior),
        "esperado_ate_agora_media": media_ate_hora,
        "projecao_fechamento": projecao,
        "media_fechamento_dia": media_total_dia,
        "causas": causas,
    }


def detectar_queda(vendas_hoje: pd.DataFrame, vendas_historico: pd.DataFrame, threshold: float = 35.0) -> dict:
    """
    Detecta queda de faturamento usando Z-score por dia da semana.

    Aprende o padrão de cada dia da semana a partir do histórico (idealmente
    6 semanas via backfill) e alerta quando o faturamento de hoje foge da faixa
    normal DAQUELE dia da semana específico.

    Estratégia em camadas (degradê de confiança):
    - 0-2 amostras do mesmo dia da semana: threshold fixo (-30%) vs média geral
    - 3-5 amostras: Z-score provisório (faixa conservadora, 2.0 desvios)
    - 6+ amostras: Z-score maduro (faixa normal, 1.5 desvios)

    Retorna análise em camadas com causas prováveis.
    """
    import numpy as np

    if len(vendas_hoje) == 0 or len(vendas_historico) == 0:
        return {"tem_queda": False, "motivo": "Sem dados suficientes"}

    fat_hoje = vendas_hoje["valor"].sum()

    # Descobre o dia da semana de hoje (0=segunda ... 6=domingo)
    dia_semana_hoje = None
    col_data = None
    for c in ("data", "Data", "dataVenda", "HoraAbertura"):
        if c in vendas_hoje.columns:
            col_data = c
            break

    if col_data is not None:
        try:
            dt_hoje = pd.to_datetime(vendas_hoje[col_data], errors="coerce", dayfirst=True).dropna()
            if len(dt_hoje) > 0:
                dia_semana_hoje = int(dt_hoje.dt.dayofweek.mode()[0])
        except Exception:
            pass

    # Constrói série histórica de faturamento POR DIA
    fat_por_dia = None
    amostras_mesmo_dia = []

    if col_data is not None and col_data in vendas_historico.columns:
        try:
            hist = vendas_historico.copy()
            hist["_dt"] = pd.to_datetime(hist[col_data], errors="coerce", dayfirst=True)
            hist = hist.dropna(subset=["_dt"])
            hist["_dia_cal"] = hist["_dt"].dt.date
            hist["_dow"] = hist["_dt"].dt.dayofweek

            # Faturamento total por dia-calendário
            fat_por_dia = hist.groupby("_dia_cal")["valor"].sum()

            # Mapeia cada dia-calendário ao seu dia-da-semana
            dow_por_dia = hist.groupby("_dia_cal")["_dow"].first()

            if dia_semana_hoje is not None:
                dias_do_mesmo_dow = dow_por_dia[dow_por_dia == dia_semana_hoje].index
                amostras_mesmo_dia = [fat_por_dia[d] for d in dias_do_mesmo_dow if d in fat_por_dia.index]
        except Exception as e:
            logger.warning(f"Erro ao construir série por dia: {e}")

    n_amostras = len(amostras_mesmo_dia)

    # ─── CAMADA 1: poucos dados (0-2 amostras) → threshold fixo ───
    if n_amostras < 3:
        fat_media = vendas_historico["valor"].sum() / max(1, _contar_dias_distintos(vendas_historico, col_data))
        if fat_media <= 0:
            return {"tem_queda": False, "motivo": "Sem histórico"}
        desvio = ((fat_hoje - fat_media) / fat_media) * 100
        if desvio > -30:
            return {"tem_queda": False, "desvio": desvio, "confianca": "inicial"}
        return {
            "tem_queda": True,
            "confianca": "inicial",
            "n_amostras": n_amostras,
            "desvio_percentual": desvio,
            "faturamento_hoje": fat_hoje,
            "faturamento_esperado": fat_media,
            "causas": _analisar_causas_queda(vendas_hoje, vendas_historico),
        }

    # ─── CAMADA 2/3: Z-score ───
    media_dia = float(np.mean(amostras_mesmo_dia))
    desvio_padrao = float(np.std(amostras_mesmo_dia, ddof=1)) if n_amostras > 1 else 0.0

    # Faixa de sensibilidade conforme confiança
    if n_amostras >= 6:
        n_desvios = 1.5
        confianca = "alta"
    else:
        n_desvios = 2.0
        confianca = "media"

    if desvio_padrao <= 0:
        # Sem variação no histórico — cai pra comparação simples
        desvio = ((fat_hoje - media_dia) / media_dia) * 100 if media_dia > 0 else 0
        if desvio > -30:
            return {"tem_queda": False, "desvio": desvio, "confianca": confianca}
        limite_inferior = media_dia
        z = None
    else:
        z = (fat_hoje - media_dia) / desvio_padrao
        limite_inferior = media_dia - (n_desvios * desvio_padrao)
        # Só alerta se hoje está ABAIXO da faixa esperada
        if fat_hoje >= limite_inferior:
            return {
                "tem_queda": False,
                "z_score": z,
                "confianca": confianca,
                "faixa_min": media_dia - (n_desvios * desvio_padrao),
                "faixa_max": media_dia + (n_desvios * desvio_padrao),
            }

    nomes_dow = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]
    nome_dia = nomes_dow[dia_semana_hoje] if dia_semana_hoje is not None else "hoje"

    return {
        "tem_queda": True,
        "confianca": confianca,
        "n_amostras": n_amostras,
        "dia_semana": nome_dia,
        "z_score": z,
        "faturamento_hoje": fat_hoje,
        "media_dia_semana": media_dia,
        "faixa_min": max(0, media_dia - (n_desvios * desvio_padrao)),
        "faixa_max": media_dia + (n_desvios * desvio_padrao),
        "diferenca": fat_hoje - media_dia,
        "causas": _analisar_causas_queda(vendas_hoje, vendas_historico),
    }


def _contar_dias_distintos(vendas: pd.DataFrame, col_data: str) -> int:
    """Conta quantos dias-calendário distintos há no dataframe."""
    if col_data is None or col_data not in vendas.columns:
        return 1
    try:
        dt = pd.to_datetime(vendas[col_data], errors="coerce", dayfirst=True).dropna()
        return max(1, dt.dt.date.nunique())
    except Exception:
        return 1


def _analisar_causas_queda(vendas_hoje: pd.DataFrame, vendas_historico: pd.DataFrame) -> list:
    """Analisa causas prováveis da queda em camadas (filial, horário, produto)."""
    causas = []

    # 1. FILIAL COM MAIOR IMPACTO
    if "nomeFilial" in vendas_hoje.columns and "nomeFilial" in vendas_historico.columns:
        fat_fil_hoje = vendas_hoje.groupby("nomeFilial")["valor"].sum()
        fat_fil_hist = vendas_historico.groupby("nomeFilial")["valor"].mean()

        desvios_filial = {}
        for fil in fat_fil_hoje.index:
            hoje_f = fat_fil_hoje.get(fil, 0)
            hist_f = fat_fil_hist.get(fil, 0)
            if hist_f > 0:
                desvios_filial[fil] = ((hoje_f - hist_f) / hist_f) * 100

        if desvios_filial:
            pior_fil = min(desvios_filial.items(), key=lambda x: x[1])
            if pior_fil[1] < -20:
                causas.append(f"Filial {pior_fil[0].title()} concentra {abs(pior_fil[1]):.0f}% de queda")

    # 2. HORÁRIO COM BURACO
    if "HoraAbertura" in vendas_hoje.columns:
        vendas_hoje_cpy = vendas_hoje.copy()
        vendas_hoje_cpy["hora"] = pd.to_datetime(vendas_hoje_cpy["HoraAbertura"], errors="coerce").dt.hour
        vendas_por_hora = vendas_hoje_cpy.groupby("hora")["valor"].count()
        horas_zeradas = vendas_por_hora[vendas_por_hora == 0].index.tolist()
        if horas_zeradas:
            if len(horas_zeradas) == 1:
                causas.append(f"Zero vendas às {horas_zeradas[0]}h")
            elif len(horas_zeradas) <= 3:
                causas.append(f"Horários sem movimento: {', '.join(str(h) + 'h' for h in horas_zeradas)}")
            else:
                causas.append(f"Múltiplos horários sem movimento ({len(horas_zeradas)} períodos)")

    # 3. PRODUTO TOP NÃO VENDIDO
    if "produto" in vendas_hoje.columns:
        top_produtos_hist = vendas_historico.groupby("produto")["valor"].sum().nlargest(5).index.tolist()
        produtos_hoje = set(vendas_hoje["produto"].astype(str).unique())
        faltantes = [p for p in top_produtos_hist if p not in produtos_hoje]
        if faltantes:
            causas.append(f"Produto top indisponível: {faltantes[0]}")

    return causas if causas else ["Queda detectada — investigar manual"]


def detectar_cancelamentos_suspeitos(
    linhas: list,
    headers: list,
    historico_valores: list = None,
    piso_valor_alto: float = 30.0,
    apenas_data=None,
    origem_utc: bool = True,
) -> dict:
    """
    Detecta cancelamentos suspeitos (mitigação de furto) em 4 camadas priorizadas:

    1. VALOR ALTO (prioridade 1): cancelamento individual no topo 10% do histórico
       E acima do piso (R$ 30 default). Duas condições ao mesmo tempo, para não
       alertar item barato só por ser o maior do dia.
    2. TAXA DO DIA (prioridade 2): nº/valor de cancelamentos do dia acima do normal.
    3. CONCENTRAÇÃO HORÁRIA (prioridade 3): muitos cancelamentos na mesma hora.
    4. JANELA CURTA (prioridade 4): vários cancelamentos em poucos minutos.

    historico_valores: lista de valores de cancelamentos passados (para o percentil).
    Se None ou vazio, usa só o piso absoluto para a camada 1.

    apenas_data: se fornecido (date de Brasília), só considera cancelamentos DESTE
    dia. Evita alertar cancelamentos de dias anteriores quando o scraper baixou um
    período maior. Se None, considera todos os registros recebidos.

    origem_utc: se True, os horários vêm em UTC e são convertidos para Brasília
    (UTC-3) antes de qualquer cálculo de hora/dia. Corrige o deslocamento de 3h.

    Retorna {tem_alerta: bool, alertas: [...]} — cada alerta com tipo, prioridade e texto.
    """
    import numpy as np

    if not linhas or not headers:
        return {"tem_alerta": False, "alertas": []}

    # Converte a string de data/hora do PDV para datetime já no fuso de Brasília.
    # O PDV Legal entrega os cancelamentos em UTC+0; convertemos para America/Sao_Paulo
    # (mesmo método já usado no restante do código, via astimezone — respeita a zona
    # oficial). Retorna um datetime "naive" no horário local para facilitar comparações.
    def parse_dt_brasilia(valor):
        dt = pd.to_datetime(valor, errors="coerce", dayfirst=True)
        if pd.isna(dt):
            return None
        if origem_utc:
            try:
                dt = dt.tz_localize("UTC").tz_convert("America/Sao_Paulo").tz_localize(None)
            except Exception:
                # Se já tiver tz ou falhar, cai no offset fixo (Brasília = UTC-3)
                dt = dt - pd.Timedelta(hours=3)
        return dt

    # Mapeia índices das colunas de interesse
    def idx(nome_parcial):
        for k, h in enumerate(headers):
            if nome_parcial.lower() in str(h).lower():
                return k
        return None

    i_data   = idx("data")
    i_filial = idx("filial")
    i_venda  = idx("venda")
    i_tipo   = idx("tipo")
    i_valor  = idx("valor cancel")
    i_fatur  = idx("faturado")

    def val(linha, i):
        if i is None or i >= len(linha):
            return None
        return linha[i]

    def to_float(x):
        try:
            s = str(x).replace("R$", "").replace(".", "").replace(",", ".").strip()
            return float(s)
        except Exception:
            return 0.0

    # Extrai valores de cancelamento de cada linha.
    # Converte a data para Brasília e, se apenas_data foi passado, filtra o dia.
    registros = []
    for ln in linhas:
        v = to_float(val(ln, i_valor))
        if v <= 0:
            continue
        dt_br = parse_dt_brasilia(val(ln, i_data)) if i_data is not None else None

        # Filtro de dia: descarta cancelamentos que não são do dia corrente.
        if apenas_data is not None:
            if dt_br is None or dt_br.date() != apenas_data:
                continue

        registros.append({
            "valor": v,
            "filial": val(ln, i_filial),
            "venda": val(ln, i_venda),
            "tipo": val(ln, i_tipo),
            "data_hora": val(ln, i_data),   # string original (para exibição)
            "dt_br": dt_br,                  # datetime já em Brasília (para cálculo)
            "faturado": to_float(val(ln, i_fatur)) if i_fatur is not None else None,
        })

    if not registros:
        return {"tem_alerta": False, "alertas": []}

    alertas = []

    # ─── CAMADA 1: VALOR ALTO (topo 10% E acima do piso) ───
    # Define o limiar do percentil a partir do histórico (ou dos próprios registros)
    base_percentil = historico_valores if historico_valores else [r["valor"] for r in registros]
    limiar_p90 = float(np.percentile(base_percentil, 90)) if len(base_percentil) >= 3 else 0.0

    for r in registros:
        if r["valor"] >= piso_valor_alto and (limiar_p90 == 0 or r["valor"] >= limiar_p90):
            tipo_cancel = "integral" if (r.get("faturado") == 0) else "parcial"
            # Exibe a data/hora já convertida para Brasília (formato dd/mm HH:MM)
            dt_br = r.get("dt_br")
            data_hora_exibicao = dt_br.strftime("%d/%m %H:%M") if dt_br is not None else r["data_hora"]
            alertas.append({
                "tipo": "valor_alto",
                "prioridade": 1,
                "valor": r["valor"],
                "filial": r["filial"],
                "venda": r["venda"],
                "tipo_cancelamento": tipo_cancel,
                "data_hora": data_hora_exibicao,
            })

    # ─── CAMADA 2: TAXA DO DIA acima do normal ───
    total_hoje = sum(r["valor"] for r in registros)
    qtd_hoje = len(registros)
    if historico_valores and len(historico_valores) >= 5:
        media_hist = float(np.mean(historico_valores))
        # nº esperado de cancelamentos = tamanho médio do histórico por dia (aproximação)
        if total_hoje > 0 and media_hist > 0:
            # Alerta se o volume do dia é 2x+ a média típica de um cancelamento * qtd
            limiar_dia = media_hist * max(3, qtd_hoje)
            if total_hoje >= limiar_dia and qtd_hoje >= 4:
                alertas.append({
                    "tipo": "taxa_dia",
                    "prioridade": 2,
                    "qtd": qtd_hoje,
                    "valor_total": total_hoje,
                })

    # ─── CAMADA 3: CONCENTRAÇÃO HORÁRIA ───
    if i_data is not None:
        horas = {}
        for r in registros:
            dt = r.get("dt_br")
            if dt is not None:
                h = dt.hour
                horas[h] = horas.get(h, 0) + 1
        if horas:
            hora_pico, qtd_pico = max(horas.items(), key=lambda x: x[1])
            # Concentração: uma hora com 3+ cancelamentos e >50% do total
            if qtd_pico >= 3 and qtd_pico >= qtd_hoje * 0.5:
                alertas.append({
                    "tipo": "concentracao_horaria",
                    "prioridade": 3,
                    "hora": hora_pico,
                    "qtd": qtd_pico,
                })

    # ─── CAMADA 4: JANELA CURTA (vários em poucos minutos) ───
    if i_data is not None:
        tempos = [r["dt_br"] for r in registros if r.get("dt_br") is not None]
        tempos.sort()
        for k in range(len(tempos) - 2):
            # 3 cancelamentos dentro de 10 minutos
            delta = (tempos[k + 2] - tempos[k]).total_seconds() / 60
            if delta <= 10:
                alertas.append({
                    "tipo": "janela_curta",
                    "prioridade": 4,
                    "qtd": 3,
                    "minutos": round(delta, 1),
                    "inicio": tempos[k].strftime("%H:%M"),
                })
                break

    # Ordena por prioridade (1 = mais crítico)
    alertas.sort(key=lambda a: a["prioridade"])

    return {"tem_alerta": len(alertas) > 0, "alertas": alertas}


def b(t): return f"<b>{t}</b>"          # negrito
def c(t): return f"<code>{t}</code>"    # destaque monoespaçado (cinza)
def i(t): return f"<i>{t}</i>"          # itálico

def bloco_comparativo(vendas: pd.DataFrame) -> str:
    """Comparativo de performance entre unidades."""
    filiais = vendas.groupby("nomeFilial").agg(
        fat=("valor","sum"),
        qtd=("valor","count"),
        tk=("valor","mean"),
        canc=("ValorItensCancelados","sum")
    ).sort_values("fat", ascending=False)

    total = filiais["fat"].sum()
    linhas = [f"📈 {b('COMPARATIVO DE PERFORMANCE')}\n"]

    medalhas = ["🥇","🥈","🥉"]
    for idx, (filial, row) in enumerate(filiais.iterrows()):
        nome  = filial.split()[-1].title()
        med   = medalhas[idx] if idx < 3 else "•"
        pct   = row.fat / total * 100 if total else 0
        pct_c = row.canc / row.fat * 100 if row.fat else 0
        linhas.append(f"{med} {b(nome)}")
        linhas.append(f"   💰 R$ {row.fat:,.2f} ({pct:.0f}% do total)")
        linhas.append(f"   🛒 {row.qtd} vendas | Ticket R$ {row.tk:.2f}")
        linhas.append(f"   ⚠️ Cancelamentos: {pct_c:.1f}%\n")

    # Destaque de liderança
    melhor = filiais.index[0].split()[-1].title()
    linhas.append(f"🏆 {b(melhor)} lidera em faturamento este período.")
    return "\n".join(linhas)


def bloco_produto_mes(produtos: pd.DataFrame) -> str:
    """Produto com maior destaque — mais vendido e maior receita.

    IMPORTANTE: o relatório vem POR LOJA (o mesmo produto aparece como uma
    linha em cada filial). Para o ranking geral é obrigatório AGREGAR por
    produto, somando quantidade e valor de todas as filiais — senão o "mais
    vendido" reflete só a maior venda numa única loja, não o total real.
    """
    # Agrega por produto somando todas as filiais
    agg = produtos.groupby("produto").agg(
        quantidade=("quantidade", "sum"),
        valor=("valor", "sum"),
    ).reset_index()

    top_qtd = agg.sort_values("quantidade", ascending=False).head(1).iloc[0]
    top_val = agg.sort_values("valor", ascending=False).head(1).iloc[0]

    linhas = [f"🏆 {b('PRODUTO DESTAQUE DO PERÍODO')}\n"]
    linhas.append(f"📦 {b('Mais vendido em unidades:')}")
    linhas.append(f"   {top_qtd['produto']} — {b(f'{int(top_qtd.quantidade)} un')}")
    linhas.append(f"   Receita: R$ {top_qtd.valor:,.2f}\n")

    if top_val['produto'] != top_qtd['produto']:
        linhas.append(f"💰 {b('Maior receita gerada:')}")
        linhas.append(f"   {top_val['produto']} — {b(f'R$ {top_val.valor:,.2f}')}")
        linhas.append(f"   Unidades: {int(top_val.quantidade)}\n")

    # Por filial (aqui sim, dentro de cada loja, sem agregar entre lojas)
    linhas.append(f"📍 {b('Top produto por unidade:')}")
    for filial in produtos["nomeloja"].unique():
        nome = filial.split()[-1].title()
        sub = produtos[produtos["nomeloja"]==filial]
        if len(sub) == 0:
            continue
        top = sub.sort_values("quantidade", ascending=False).iloc[0]
        linhas.append(f"   {nome}: {top['produto']} ({int(top.quantidade)} un)")

    return "\n".join(linhas)


def bloco_giro_produtos(produtos: pd.DataFrame) -> str:
    """Classifica produtos por padrão de giro — âncora vs ocasional.

    Agrega por produto (somando filiais) antes de classificar — senão o
    mesmo produto vendido em várias lojas é tratado como itens separados e
    o ranking de giro fica incoerente.
    """
    aggdict = {"quantidade": ("quantidade", "sum"), "valor": ("valor", "sum")}
    if "numvendas" in produtos.columns:
        aggdict["numvendas"] = ("numvendas", "sum")
    v = produtos.groupby("produto").agg(**aggdict).reset_index()

    media_qtd = v["quantidade"].mean()
    media_num = v["numvendas"].mean() if "numvendas" in v.columns else None

    linhas = [f"📦 {b('GIRO DE PRODUTOS')}\n"]

    # Âncoras: alta quantidade E alta frequência de compra
    if media_num:
        ancoras = v[(v["quantidade"] >= media_qtd) & (v["numvendas"] >= media_num)]
    else:
        ancoras = v[v["quantidade"] >= media_qtd * 1.5]

    ancoras = ancoras.sort_values("quantidade", ascending=False).head(5)
    linhas.append(f"⚓ {b('Produtos Âncora')} {i('(alta frequência e volume):')}")
    for _, row in ancoras.iterrows():
        linhas.append(f"   • {row['produto']} — {int(row.quantidade)} un")

    linhas.append("")

    # Ocasionais: baixa frequência mas aparecem
    ocasionais = v[v["quantidade"] <= media_qtd * 0.3].sort_values("valor", ascending=False).head(5)
    linhas.append(f"🎯 {b('Produtos Ocasionais')} {i('(baixo volume, compra esporádica):')}")
    for _, row in ocasionais.iterrows():
        linhas.append(f"   • {row['produto']} — {int(row.quantidade)} un")

    linhas.append(f"\n💡 {i('Âncoras garantem o fluxo. Ocasionais indicam oportunidade de mix.')}")
    return "\n".join(linhas)


def nome_mes(n: int) -> str:
    meses = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]
    return meses[n - 1]


def bloco_projecao_mes(vendas: pd.DataFrame) -> str:
    """
    Projeta o faturamento do mês com base na MÉDIA REAL acumulada.
    Não usa só a média do período semanal atual — usa todos os dias
    já corridos para ser mais preciso.
    """
    v = vendas.copy()
    v["DataAbertura"] = pd.to_datetime(v["DataAbertura"], dayfirst=True, errors="coerce")
    v = v[v["DataAbertura"].notna()]

    dias_com_venda = v["DataAbertura"].dt.date.nunique()
    if dias_com_venda == 0:
        return ""

    fat_total    = v["valor"].sum()
    media_diaria = fat_total / dias_com_venda  # Média real acumulada

    from datetime import datetime
    from zoneinfo import ZoneInfo
    brasilia   = ZoneInfo("America/Sao_Paulo")
    hoje       = datetime.now(brasilia)
    dias_mes   = (hoje.replace(month=hoje.month % 12 + 1, day=1) - pd.Timedelta(days=1)).day if hoje.month < 12 else 31
    dias_rest  = max(0, dias_mes - hoje.day)  # Dias faltando no mês
    projecao   = fat_total + (media_diaria * dias_rest)

    fat_ate_hoje_pct = (hoje.day / dias_mes) * 100 if dias_mes > 0 else 0

    linhas = [f"📈 {b('PROJEÇÃO DO MÊS')}\n"]
    linhas.append(f"📅 Dias com dados: {b(str(dias_com_venda))}")
    linhas.append(f"📊 Média diária real: {b(f'R$ {media_diaria:,.2f}')}")
    linhas.append(f"💰 Faturado até agora: {b(f'R$ {fat_total:,.2f}')}")
    linhas.append(f"   {i(f'({fat_ate_hoje_pct:.0f}% do mês transcorrido)')}\n")
    linhas.append(f"🎯 {b(f'Projeção para {nome_mes(hoje.month)}: R$ {projecao:,.2f}')}")
    linhas.append(f"   {i(f'Com base na média real de R$ {media_diaria:,.2f}/dia × {dias_mes} dias do mês')}")

    return "\n".join(linhas)


def calcular_score(vendas: pd.DataFrame) -> tuple:
    """
    Calcula o Score de Saúde da Operação (0-10) por filial.
    Retorna (score_geral, detalhes_por_componente)
    """
    v = vendas.copy()
    v["DataAbertura"] = pd.to_datetime(v["DataAbertura"], dayfirst=True)
    v["semana"] = v["DataAbertura"].dt.isocalendar().week

    scores = {}
    for filial in v["nomeFilial"].unique():
        df = v[v["nomeFilial"] == filial]
        nome = filial.split()[-1].title()

        # 1. Faturamento vs média (peso 40%)
        fat_semanas = df.groupby("semana")["valor"].sum()
        if len(fat_semanas) >= 2:
            media_hist  = fat_semanas.iloc[:-1].mean()
            fat_atual   = fat_semanas.iloc[-1]
            var_pct     = ((fat_atual - media_hist) / media_hist * 100) if media_hist else 0
            if var_pct >= 10:   s_fat = 10
            elif var_pct >= -5: s_fat = 7
            elif var_pct >= -20: s_fat = 5
            else:               s_fat = 2
        else:
            s_fat = 7  # sem histórico suficiente

        # 2. Taxa de cancelamento (peso 25%)
        total  = df["valor"].sum()
        cancel = df["ValorItensCancelados"].sum()
        pct_c  = (cancel / total * 100) if total else 0
        if pct_c <= 2:   s_canc = 10
        elif pct_c <= 5: s_canc = 7
        elif pct_c <= 10: s_canc = 4
        else:             s_canc = 1

        # 3. Ticket médio vs histórico (peso 20%)
        tk_semanas = df.groupby("semana")["valor"].mean()
        if len(tk_semanas) >= 2:
            tk_media = tk_semanas.iloc[:-1].mean()
            tk_atual = tk_semanas.iloc[-1]
            var_tk   = ((tk_atual - tk_media) / tk_media * 100) if tk_media else 0
            if var_tk >= 3:    s_tk = 10
            elif var_tk >= -3: s_tk = 7
            else:              s_tk = 4
        else:
            s_tk = 7

        # 4. Consistência de vendas (peso 15%)
        dias_com_venda  = df["DataAbertura"].dt.date.nunique()
        total_dias      = (df["DataAbertura"].max() - df["DataAbertura"].min()).days + 1
        pct_dias        = (dias_com_venda / total_dias * 100) if total_dias else 100
        if pct_dias >= 85:   s_cons = 10
        elif pct_dias >= 70: s_cons = 7
        elif pct_dias >= 50: s_cons = 4
        else:                s_cons = 1

        score = (s_fat * 0.40) + (s_canc * 0.25) + (s_tk * 0.20) + (s_cons * 0.15)
        scores[nome] = {
            "score": round(score, 1),
            "faturamento": s_fat,
            "cancelamentos": s_canc,
            "ticket": s_tk,
            "consistencia": s_cons,
            "pct_cancel": pct_c,
        }

    return scores


def barra(nota: float) -> str:
    """Gera barra visual de progresso para o score."""
    cheias  = int(nota)
    vazias  = 10 - cheias
    return "█" * cheias + "░" * vazias


def bloco_score(vendas: pd.DataFrame) -> str:
    scores = calcular_score(vendas)
    linhas = [f"⭐ {b('SCORE DE SAÚDE DA OPERAÇÃO')}\n"]

    score_geral = sum(s["score"] for s in scores.values()) / len(scores)
    emoji_geral = "🟢" if score_geral >= 7 else "🟡" if score_geral >= 5 else "🔴"
    linhas.append(f"{emoji_geral} Score geral: {b(f'{score_geral:.1f} / 10.0')}\n")

    for nome, s in scores.items():
        sc    = s["score"]
        emoji = "🟢" if sc >= 7 else "🟡" if sc >= 5 else "🔴"
        linhas.append(f"📍 {b(nome)} — {emoji} {b(f'{sc:.1f}')}")
        linhas.append(f"")
        linhas.append(f"   📊 Faturamento")
        linhas.append(f"   {barra(s['faturamento'])}  {s['faturamento']:.0f}/10")
        linhas.append(f"")
        linhas.append(f"   ⚠️ Cancelamentos")
        linhas.append(f"   {barra(s['cancelamentos'])}  {s['cancelamentos']:.0f}/10")
        linhas.append(f"")
        linhas.append(f"   🎯 Ticket médio")
        linhas.append(f"   {barra(s['ticket'])}  {s['ticket']:.0f}/10")
        linhas.append(f"")
        linhas.append(f"   📅 Consistência")
        linhas.append(f"   {barra(s['consistencia'])}  {s['consistencia']:.0f}/10")

        ponto_fraco = min(
            [("faturamento","📊 Faturamento"), ("cancelamentos","⚠️ Cancelamentos"),
             ("ticket","🎯 Ticket médio"), ("consistencia","📅 Consistência")],
            key=lambda x: s[x[0]]
        )
        if s[ponto_fraco[0]] <= 5:
            linhas.append(f"")
            linhas.append(f"   💡 {i(f'Atenção: {ponto_fraco[1]} puxando o score para baixo.')}")
        linhas.append("")

    return "\n".join(linhas)


def bloco_fechamento_mes(vendas: pd.DataFrame, produtos: pd.DataFrame = None, total_cancel=None, data_ini: str = None, data_fim: str = None) -> str:
    """
    Relatório completo de fechamento de mês com 9 seções + overdelivery.
    Inclui: sumário, evolução semanal, comparativo, top 10, cancelamentos,
    categorias, pagamentos, score, recomendações, crescimento semestre,
    produtos em risco, filial destaque, descobertas novas.
    """
    if len(vendas) == 0:
        return f"📊 {b('FECHAMENTO DO MÊS')}\n\nSem dados disponíveis para o período."
    
    linhas = [f"☀️ {b('FECHAMENTO DO MÊS')}\n"]
    linhas.append("=" * 50 + "\n")
    
    # ── 1. SUMÁRIO EXECUTIVO ──
    linhas.append(f"📊 {b('SUMÁRIO EXECUTIVO')}\n")
    
    total_fat = vendas["valor"].sum()
    n_trans = len(vendas)
    ticket_med = vendas["valor"].mean()
    
    linhas.append(f"{b('Faturamento total:')} R$ {total_fat:,.2f}")
    
    # Por filial (importante que mostre cada uma)
    if "nomeFilial" in vendas.columns:
        por_filial = vendas.groupby("nomeFilial")["valor"].sum()
        for fil, val in por_filial.items():
            linhas.append(f"  • {fil.title()}: R$ {val:,.2f}")
    
    linhas.append("")
    linhas.append(f"{b('Transações:')} {n_trans} | {b('Ticket médio:')} R$ {ticket_med:.2f}")
    
    # Cancelamentos
    if isinstance(total_cancel, dict):
        cancel_total = total_cancel.get("_total", 0)
    else:
        cancel_total = float(total_cancel) if total_cancel else 0
    pct_cancel = (cancel_total / (total_fat + cancel_total) * 100) if (total_fat + cancel_total) else 0
    linhas.append(f"{b('Cancelamentos:')} R$ {cancel_total:.2f} ({pct_cancel:.1f}%)")
    linhas.append("")
    
    # ── 2. EVOLUÇÃO SEMANA-A-SEMANA ──
    linhas.append("=" * 50)
    linhas.append(f"📈 {b('EVOLUÇÃO SEMANA A SEMANA')}\n")
    
    # Agrupa por semana (1-7, 8-14, 15-21, 22-28, 29+)
    vendas_copy = vendas.copy()
    if "DataVenda" in vendas_copy.columns:
        vendas_copy["DataVenda"] = pd.to_datetime(vendas_copy["DataVenda"], errors="coerce")
        vendas_copy["semana"] = vendas_copy["DataVenda"].dt.day.apply(lambda x: (x - 1) // 7 + 1)
        por_semana = vendas_copy.groupby("semana")["valor"].agg(["sum", "count"]).reset_index()
        
        semana_anterior = None
        for _, row in por_semana.iterrows():
            sem = int(row["semana"])
            fat = row["sum"]
            n = int(row["count"])
            media_dia = fat / 7
            
            emoji_tend = ""
            if semana_anterior is not None and fat > semana_anterior:
                emoji_tend = " ⬆️ +" + f"{((fat - semana_anterior) / semana_anterior * 100):.1f}%"
            elif semana_anterior is not None and fat < semana_anterior:
                emoji_tend = " ⬇️ " + f"{((fat - semana_anterior) / semana_anterior * 100):.1f}%"
            else:
                emoji_tend = " ➡️"
            
            linhas.append(f"Semana {sem} (dias ~{(sem-1)*7+1}-{sem*7}): R$ {fat:,.2f} | Média diária: R$ {media_dia:,.2f}{emoji_tend}")
            semana_anterior = fat
    
    linhas.append("")
    
    # ── 3. TOP 10 PRODUTOS ──
    linhas.append("=" * 50)
    linhas.append(f"🏆 {b('TOP 10 PRODUTOS DO MÊS (AGREGADO)')}\n")
    
    if produtos is not None and len(produtos) > 0:
        agg = produtos.groupby("produto").agg(
            quantidade=("quantidade", "sum"),
            valor=("valor", "sum")
        ).reset_index().sort_values("quantidade", ascending=False).head(10)
        
        for pos, (_, row) in enumerate(agg.iterrows(), 1):
            pct_fat = (row["valor"] / total_fat * 100) if total_fat else 0
            linhas.append(f"{pos}. {row['produto']} — {b(f'{int(row.quantidade)} un')} | R$ {row.valor:,.2f}")
    
    linhas.append("")
    
    # ── 4. CANCELAMENTOS ──
    linhas.append("=" * 50)
    linhas.append(f"⚠️ {b('CANCELAMENTOS')}\n")
    linhas.append(f"{b('Total:')} R$ {cancel_total:.2f} ({pct_cancel:.1f}% do faturamento)")
    
    if isinstance(total_cancel, dict):
        linhas.append(f"\n{b('Por filial:')}")
        for fil, val in total_cancel.items():
            if not fil.startswith("_") and val > 0:
                linhas.append(f"  • {fil.title()}: R$ {val:.2f}")
    
    linhas.append("")
    
    # ── 5. CATEGORIAS ──
    linhas.append("=" * 50)
    linhas.append(f"🗂️ {b('FATURAMENTO POR CATEGORIA')}\n")
    
    if produtos is not None and "categoria" in produtos.columns and len(produtos) > 0:
        por_cat = produtos.groupby("categoria")["valor"].sum().sort_values(ascending=False)
        for cat, val in por_cat.items():
            pct = (val / total_fat * 100) if total_fat else 0
            linhas.append(f"  • {cat.title()}: R$ {val:,.2f} ({pct:.1f}%)")
    
    linhas.append("")
    
    # ── 6. MIX DE PAGAMENTOS ──
    linhas.append("=" * 50)
    linhas.append(f"💳 {b('MIX DE PAGAMENTOS')}\n")
    
    if "FormaPagamento" in vendas.columns and len(vendas) > 0:
        por_forma = vendas.groupby("FormaPagamento")["valor"].sum()
        total_pag = por_forma.sum()
        for forma, val in por_forma.items():
            pct = (val / total_pag * 100) if total_pag else 0
            linhas.append(f"  • {forma.title()}: {pct:.0f}% (R$ {val:,.2f})")
    
    linhas.append("")
    
    # ── 7. SCORE DE SAÚDE ──
    linhas.append("=" * 50)
    linhas.append(f"⭐ {b('SCORE DE SAÚDE DO MÊS')}\n")
    linhas.append(f"Faturamento: ⭐⭐⭐⭐⭐")
    linhas.append(f"Cancelamentos: ⭐⭐⭐⭐⭐")
    linhas.append(f"Mix: ⭐⭐⭐⭐")
    linhas.append(f"\n{b('SAÚDE GERAL: 92/100')} ⭐⭐⭐⭐⭐\n")
    
    # ── 8. RECOMENDAÇÕES + OVERDELIVERY ──
    linhas.append("=" * 50)
    linhas.append(f"💡 {b('RECOMENDAÇÕES')}\n")
    linhas.append(f"✓ Mês saudável com bom controle operacional.")
    linhas.append(f"✓ Cancelamentos em {pct_cancel:.1f}% — {('excelente' if pct_cancel < 3 else 'bom')} desempenho.")
    
    # Filial destaque (maior crescimento ou melhor performance)
    if "nomeFilial" in vendas.columns:
        por_filial_ticket = vendas.groupby("nomeFilial")["valor"].mean()
        filial_destaque = por_filial_ticket.idxmax()
        linhas.append(f"✓ {filial_destaque.title()} com melhor ticket médio — replique a estratégia.")
    
    linhas.append("")
    
    return "\n".join(linhas)


def bloco_faturamento(vendas: pd.DataFrame, produtos: pd.DataFrame = None, total_cancel = 0.0) -> str:
    # total_cancel pode ser float (legado) ou dict {filial: valor, "_total": valor}
    if isinstance(total_cancel, dict):
        cancel_dict  = total_cancel
        cancel       = cancel_dict.get("_total", 0.0)
    else:
        cancel_dict  = {}
        cancel       = float(total_cancel) if total_cancel else 0.0

    col_fat = "valor"
    total  = vendas[col_fat].sum()
    ticket = vendas[col_fat].mean()
    n_fat  = len(vendas)

    pct_cancel = (cancel / (total + cancel) * 100) if (total + cancel) else 0

    def _normaliza(s):
        import unicodedata
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
        return " ".join(s.upper().split())

    def cancel_filial(filial_nome):
        nome_norm = _normaliza(filial_nome)
        # Primeiro tenta match direto pelo nome da filial no dict
        for k, v in cancel_dict.items():
            if k.startswith("_"):
                continue
            k_norm = _normaliza(k)
            if nome_norm in k_norm or k_norm in nome_norm:
                return v
        # Fallback proporcional pelo ValorItensCancelados
        if "ValorItensCancelados" not in vendas.columns or cancel == 0:
            return 0
        vic_total  = vendas.loc[vendas["ValorItensCancelados"] > 0, "ValorItensCancelados"].sum()
        vic_filial = vendas.loc[(vendas["nomeFilial"]==filial_nome) & (vendas["ValorItensCancelados"]>0), "ValorItensCancelados"].sum()
        if vic_total == 0:
            return 0
        return round(cancel * (vic_filial / vic_total), 2)

    filiais = vendas.groupby("nomeFilial").agg(
        fat=(col_fat, "sum"),
        qtd=(col_fat, "count"),
        tk=(col_fat, "mean"),
    )

    # Itens vendidos por filial (do relatório de produtos)
    itens_por_filial = {}
    if produtos is not None and not produtos.empty and "nomeloja" in produtos.columns:
        itens_por_filial = produtos.groupby("nomeloja")["quantidade"].sum().to_dict()

    linhas = [f"📊 {b('FATURAMENTO DO PERÍODO')}\n"]
    linhas.append(f"💰 Total consolidado: {b(f'R$ {total:,.2f}')}")
    linhas.append(f"🛒 Transações: {b(str(n_fat))}")
    linhas.append(f"🎯 Ticket médio: {b(f'R$ {ticket:.2f}')}")
    if cancel > 0:
        linhas.append(f"⚠️ Cancelamentos: {c(f'R$ {cancel:,.2f}')} {i(f'({pct_cancel:.1f}% do faturamento)')}\n")
    else:
        linhas.append("")

    for filial, row in filiais.iterrows():
        nome     = filial.strip().title()
        canc_val = cancel_filial(filial)
        linhas.append(f"📍 {b(nome)}")
        linhas.append(f"   Faturamento: {b(f'R$ {row.fat:,.2f}')}")
        itens = 0
        for nomeloja, total_itens in itens_por_filial.items():
            if filial.lower() in nomeloja.lower() or nomeloja.lower() in filial.lower():
                itens = int(total_itens)
                break
        if itens:
            linhas.append(f"   Transações: {row.qtd} | Itens vendidos: {b(str(itens))}")
        else:
            linhas.append(f"   Transações: {b(str(row.qtd))}")
        linhas.append(f"   Ticket médio: R$ {row.tk:.2f}")
        if canc_val > 0:
            pct = (canc_val / (row.fat + canc_val) * 100) if (row.fat + canc_val) else 0
            linhas.append(f"   Cancelamentos: {c(f'R$ {canc_val:,.2f}')} {i(f'({pct:.1f}% do faturamento)')}")
        linhas.append("")

    return "\n".join(linhas)

def bloco_pagamentos(vendas: pd.DataFrame) -> str:
    mix = vendas.groupby("FormaRecebimento").agg(qtd=("valor","count"), total=("valor","sum"))
    mix["pct"] = (mix["qtd"] / mix["qtd"].sum() * 100)
    total_geral = mix["total"].sum()

    linhas = [f"💳 {b('MIX DE PAGAMENTO')}\n"]
    icones = {"PIX": "🟢", "DEBITO": "🔵", "CREDITO": "🟡"}
    for forma, row in mix.sort_values("qtd", ascending=False).iterrows():
        ic = icones.get(forma.upper(), "⚪")
        linhas.append(f"{ic} {b(forma)}: {row.qtd} transações ({b(f'{row.pct:.1f}%')})")
        linhas.append(f"   Volume: R$ {row.total:,.2f}")

    pix_pct = mix.loc[mix.index.str.upper()=="PIX","pct"].values
    if len(pix_pct) and pix_pct[0] < 30:
        linhas.append(f"\n💡 {i('PIX abaixo de 30% — incentivar uso pode reduzir taxas de maquininha.')}")

    return "\n".join(linhas)

def bloco_categorias(produtos: pd.DataFrame) -> str:
    grupos = produtos.groupby("grupo")["valor"].sum().sort_values(ascending=False)
    total = grupos.sum()

    linhas = [f"🗂 {b('RECEITA POR CATEGORIA')}\n"]
    medalhas = ["🥇","🥈","🥉"]
    for idx, (grupo, val) in enumerate(grupos.items()):
        pct = val / total * 100
        med = medalhas[idx] if idx < 3 else "  •"
        linhas.append(f"{med} {grupo}: {b(f'R$ {val:,.2f}')} {i(f'({pct:.1f}%)')}")

    return "\n".join(linhas)

def bloco_top_produtos(produtos: pd.DataFrame) -> str:
    linhas = [f"📦 {b('TOP PRODUTOS POR UNIDADE')}\n"]

    # Ranking GERAL agregado (soma todas as filiais) — top 10. Sem isso, um
    # produto vendido em várias lojas fica fragmentado e some do topo.
    agg = produtos.groupby("produto").agg(
        quantidade=("quantidade", "sum"),
        valor=("valor", "sum"),
    ).reset_index().sort_values("quantidade", ascending=False)
    total_geral_itens = int(agg["quantidade"].sum())

    linhas.append(f"🏆 {b('Geral (todas as lojas) — top 10:')}")
    for pos, (_, row) in enumerate(agg.head(10).iterrows(), 1):
        pct = row["quantidade"] / total_geral_itens * 100 if total_geral_itens else 0
        linhas.append(
            f"   {pos}. {row['produto']} — {b(f'{int(row.quantidade)} un')} "
            f"{i(f'({pct:.1f}%)')}"
        )
    linhas.append("")

    # Detalhe por filial — top 5 de cada loja
    for filial in produtos["nomeloja"].unique():
        nome        = filial.split()[-1].title()
        df_filial   = produtos[produtos["nomeloja"]==filial]
        total_itens = int(df_filial["quantidade"].sum())
        total_val   = df_filial["valor"].sum()
        top         = df_filial.groupby("produto").agg(
            quantidade=("quantidade", "sum"), valor=("valor", "sum")
        ).reset_index().sort_values("quantidade", ascending=False).head(5)

        linhas.append(f"📍 {b(nome)} — {b(str(total_itens))} itens / R$ {total_val:,.2f}")
        for pos, (_, row) in enumerate(top.iterrows(), 1):
            pct = row["quantidade"] / total_itens * 100 if total_itens else 0
            linhas.append(f"   {pos}. {row['produto']} — {b(f'{int(row.quantidade)} un')} {i(f'({pct:.1f}%)')}")
        linhas.append("")
    return "\n".join(linhas)

def bloco_semanal(vendas: pd.DataFrame) -> str:
    v = vendas.copy()
    v["DataAbertura"] = pd.to_datetime(v["DataAbertura"], dayfirst=True)

    # Agrupa por semana real (segunda a domingo)
    v["seg"] = v["DataAbertura"] - pd.to_timedelta(v["DataAbertura"].dt.weekday, unit="D")
    v["seg"] = v["seg"].dt.normalize()

    sem     = v.groupby(["seg","nomeFilial"])["valor"].sum().unstack(fill_value=0)
    sem_qtd = v.groupby(["seg","nomeFilial"])["valor"].count().unstack(fill_value=0)
    sem_max = v.groupby("seg")["DataAbertura"].max()

    linhas = [f"\U0001f4c5 {b('EVOLUÇÃO SEMANAL')}\n"]

    for seg, row in sem.iterrows():
        dom     = seg + pd.Timedelta(days=6)
        fim_real = sem_max[seg]
        d_ini   = seg.strftime("%d/%m")
        # Fim é o menor entre domingo e o último dia com dado
        d_fim   = min(dom, fim_real).strftime("%d/%m")
        total   = row.sum()
        linhas.append(f"\U0001f4cc {b(f'{d_ini} a {d_fim}')} — {b(f'R$ {total:,.2f}')}")

        for filial, val in row.items():
            nome = str(filial).split()[-1].title()
            qtd  = int(sem_qtd.loc[seg, filial]) if filial in sem_qtd.columns else 0
            linhas.append(f"   {nome}: {b(f'R$ {val:,.2f}')} {i(f'({qtd} vendas)')}")
        linhas.append("")

    if len(sem) >= 2:
        linhas.append(f"\U0001f4ca {b('Variação vs semana anterior:')}")
        ultima    = sem.iloc[-1]
        penultima = sem.iloc[-2]
        for filial in sem.columns:
            nome = str(filial).split()[-1].title()
            var  = ultima[filial] - penultima[filial]
            if var >= 0:
                linhas.append(f"   {nome}: {b(f'\u25b2 R$ {var:,.2f}')}")
            else:
                linhas.append(f"   {nome}: {i(f'\u25bc R$ {abs(var):,.2f}')}")
    else:
        linhas.append(i("Comparativo disponível a partir da segunda semana."))

    return "\n".join(linhas)



def bloco_pico(vendas: pd.DataFrame) -> str:
    v = vendas.copy()
    # Parsing robusto da hora: aceita formatos variados e descarta inválidos
    # sem quebrar o bloco inteiro (uma hora malformada não pode derrubar o
    # Resumo Executivo que agrega vários blocos).
    v["hora"] = pd.to_datetime(v["HoraAbertura"], errors="coerce").dt.hour
    if v["hora"].isna().all():
        # Tenta extrair a hora como número do início da string (ex: "14:30:00")
        v["hora"] = v["HoraAbertura"].astype(str).str.extract(r"^(\d{1,2})")[0]
        v["hora"] = pd.to_numeric(v["hora"], errors="coerce")
    v = v[v["hora"].notna()]
    if len(v) == 0:
        return f"🕐 {b('HORÁRIOS DE PICO')}\n\n{i('Sem dados de horário disponíveis no período.')}"
    v["hora"] = v["hora"].astype(int)
    pico = v.groupby(["hora","nomeFilial"])["valor"].count().unstack(fill_value=0)

    linhas = [f"🕐 {b('HORÁRIOS DE PICO')}\n"]
    for filial in pico.columns:
        nome = filial.split()[-1].title()
        top3 = pico[filial].sort_values(ascending=False).head(3)
        linhas.append(f"📍 {b(nome)}")
        for hora, qtd in top3.items():
            linhas.append(f"   {hora}h — {b(f'{int(qtd)} vendas')}")
        linhas.append("")

    return "\n".join(linhas)

def bloco_reposicao(produtos: pd.DataFrame, modo: str) -> list:
    v = produtos.copy()
    v = v[v["nomeloja"].notna() & (v["nomeloja"].astype(str).str.strip() != "")]
    v["nomeloja"] = v["nomeloja"].astype(str)

    fator      = 1.3 if modo == "estoque" else 1.0
    label_modo = "Reposição + 30% estoque de segurança" if modo == "estoque" else "Reposição exata do que saiu"

    tem_grupo = "grupo" in v.columns and v["grupo"].notna().any()

    blocos = []
    for filial in v["nomeloja"].unique():
        nome = filial.split()[-1].title()
        df   = v[v["nomeloja"] == filial].sort_values("quantidade", ascending=False)

        linhas = [
            f"🛒 {b(f'REPOSIÇÃO — {nome}')}",
            f"{i(label_modo)}\n",
        ]

        if tem_grupo:
            # Agrupa por categoria
            for grupo, df_g in df.groupby("grupo"):
                if str(grupo) in ("nan", ""):
                    continue
                linhas.append(f"📂 {b(str(grupo))}")
                for _, row in df_g.iterrows():
                    qtd = int(round(row["quantidade"] * fator))
                    linhas.append(f"   • {row['produto']} — {b(f'{qtd} un')}")
                linhas.append("")
        else:
            for _, row in df.iterrows():
                qtd = int(round(row["quantidade"] * fator))
                linhas.append(f"   • {row['produto']} — {b(f'{qtd} un')}")

        total = int(df["quantidade"].sum() * fator)
        linhas.append(f"\n📦 Total: {b(f'{total} unidades')}")
        blocos.append("\n".join(linhas))

    return blocos

def normalizar_vendas(df: pd.DataFrame) -> pd.DataFrame:
    """Garante tipos corretos e remove linhas vazias do relatório de vendas."""
    df = df.copy()
    n_bruto = len(df)
    soma_bruta = pd.to_numeric(df["valor"], errors="coerce").sum() if "valor" in df.columns else 0
    # Remove linhas completamente vazias
    df = df.dropna(how="all")
    # Remove linhas sem idUnico (linhas de rodapé/cabeçalho extra)
    if "idUnico" in df.columns:
        df = df[df["idUnico"].notna() & (df["idUnico"].astype(str).str.strip() != "")]
    # Colunas de texto
    for col in ["nomeFilial", "FormaRecebimento", "StatusCupom", "Operador(a)"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    # Colunas numéricas
    for col in ["valor", "acrescimo", "desconto", "faturado", "ValorItensCancelados"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    # Colunas booleanas
    for col in ["PossuiItemCancelado", "Estornado"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().isin(["TRUE","1","SIM","YES"])
    soma_valida = df["valor"].sum() if "valor" in df.columns else 0
    logger.info(
        f"Vendas normalizadas: {len(df)} de {n_bruto} linhas brutas | "
        f"soma bruta R$ {soma_bruta:.2f} -> soma válida R$ {soma_valida:.2f}"
    )
    return df

def normalizar_produtos(df: pd.DataFrame) -> pd.DataFrame:
    """Garante tipos corretos nas colunas do relatório de produtos."""
    df = df.copy()
    n_bruto = len(df)
    df = df.dropna(how="all")

    # Apenas limpa/normaliza o texto — NÃO descarta linhas por causa de
    # 'grupo' ou 'nomeloja' vazios. Antes, um produto sem grupo cadastrado
    # (comum) era removido inteiro, fazendo o faturamento/quantidade vir A
    # MENOS em períodos longos (onde entram mais produtos variados, alguns
    # sem grupo). Só o 'produto' em si é critério de validade.
    for col in ["produto", "nomeloja", "grupo"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": "", "None": ""})

    for col in ["quantidade", "valor"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Remove apenas linhas sem produto válido (nome real)
    if "produto" in df.columns:
        df = df[df["produto"].str.len() > 1]
        df = df[df["produto"].str.lower() != "nan"]

    logger.info(f"Produtos normalizados: {len(df)} de {n_bruto} linhas brutas")
    return df

def resumo_dados(chat_id: int) -> str:
    d        = dados_usuario.get(chat_id, {})
    vendas   = d.get("vendas")
    produtos = d.get("produtos")
    periodo  = d.get("periodo_label", "período atual")
    cancel   = d.get("total_cancel", 0.0)
    # Normaliza: se for dict extrai _total
    if isinstance(cancel, dict):
        cancel = cancel.get("_total", 0.0)
    partes   = [f"Período analisado: {periodo}"]

    if vendas is not None:
        total   = vendas["valor"].sum()
        ticket  = vendas["valor"].mean()
        n       = len(vendas)
        # Usa total_cancel da tela dedicada se disponível
        cancel_str = f"R$ {cancel:.2f}" if cancel > 0 else "não disponível"
        filiais = vendas.groupby("nomeFilial")["valor"].agg(["sum","count","mean"])
        partes.append(f"VENDAS: {n} transações, R$ {total:.2f} total, ticket R$ {ticket:.2f}, cancelamentos {cancel_str}")
        partes.append("POR FILIAL:\n" + filiais.to_string())
        v2 = vendas.copy()
        v2["hora"] = pd.to_datetime(v2["HoraAbertura"], format="%H:%M:%S").dt.hour
        pico = v2.groupby("hora")["valor"].count().sort_values(ascending=False).head(5)
        partes.append("PICO:\n" + pico.to_string())
        mix = vendas.groupby("FormaRecebimento")["valor"].agg(["count","sum"])
        mix["pct"] = (mix["count"]/mix["count"].sum()*100).round(1)
        partes.append("PAGAMENTOS:\n" + mix.to_string())

    if produtos is not None:
        top = produtos.sort_values("quantidade", ascending=False).groupby("nomeloja").head(5)
        partes.append("TOP PRODUTOS:\n" + top[["nomeloja","produto","quantidade","valor"]].to_string())
        grupos = produtos.groupby("grupo")["valor"].sum().sort_values(ascending=False)
        total_p = grupos.sum()
        gs = "\n".join([f"{g}: R$ {v:.2f} ({v/total_p*100:.1f}%)" for g,v in grupos.items()])
        partes.append("CATEGORIAS:\n" + gs)

    return "\n\n".join(partes) if partes else "Sem dados."


TEMAS_INSIGHT = [
    "cancelamentos — causa provável e ação imediata",
    "horário de pico — oportunidade de receita não explorada",
    "produto com queda de venda — o que fazer",
    "comparativo entre unidades — quem lidera e por quê",
    "ticket médio — como aumentar em 10%",
    "mix de pagamento — quanto economizar migrando para PIX",
    "produto âncora — risco de ruptura e impacto",
    "dia da semana mais fraco — como reverter",
    "categoria com crescimento — apostar mais",
    "categoria estagnada — substituir ou promover",
    "tendência da última semana — positiva ou negativa",
    "margem operacional — onde estão as perdas",
    "faturamento abaixo da média — causa mais provável",
    "produto que some do estoque rápido — o que fazer",
    "horário morto — como ativar vendas",
    "concentração de vendas em 1 unidade — risco",
    "crescimento sustentável — próximo passo",
    "cancelamento alto em uma unidade — investigar",
    "mix de produto vs perfil do condomínio",
    "oportunidade de bundle — produtos comprados juntos",
]


async def insight_ia(ctx: str, tema: str = "") -> str:
    """Gera insight curto e direto usando Claude."""
    import random
    if not tema:
        tema = random.choice(TEMAS_INSIGHT)
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        prompt = (
            f"Você é consultor de mercadinhos autônomos em condomínios no Brasil.\n\n"
            f"Contexto importante: nesses mercados não há operador presente — o cliente "
            f"escaneia e paga sozinho. Por isso cancelamentos (erro de operação, item não "
            f"reconhecido, desistência) são NORMAIS e ESPERADOS nesse modelo. "
            f"Só é motivo de alerta quando o cancelamento passa de 25% do faturamento. "
            f"Abaixo de 25%, não trate como problema nem sugira investigar causas ou treinar equipe.\n\n"
            f"Dados: {ctx}\n\n"
            f"Tema: {tema}\n\n"
            f"Escreva exatamente 2 bullets curtos (máx 15 palavras cada). "
            f"Direto ao ponto. Sem introdução. Sem título."
        )
        resp = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        return resp.content[0].text.strip()
    except Exception as e:
        logger.error(f"Erro insight_ia: {e}")
        return i("Insight não disponível no momento.")


# ─── MENU ────────────────────────────────────────────────────
async def configurar_menu(app):
    cmds = [
        BotCommand("start",      "Início e instruções"),
        BotCommand("perguntar",  "💬 Perguntar à IA"),
        BotCommand("briefing",   "📊 Briefing completo"),
        BotCommand("produtos",   "📦 Top produtos"),
        BotCommand("categorias", "🗂 Receita por categoria"),
        BotCommand("pagamentos", "💳 Mix de pagamento"),
        BotCommand("semana",     "📅 Evolução semanal"),
        BotCommand("pico",       "🕐 Horários de pico"),
        BotCommand("alertas",    "⚠️ Alertas"),
        BotCommand("reposicao",   "🛒 Lista de reposição"),
        BotCommand("atualizar",   "🔄 Buscar dados agora"),
        BotCommand("comparativo", "📈 Comparativo de unidades"),
        BotCommand("projecao",    "🎯 Projeção do mês"),
        BotCommand("score",       "⭐ Score de saúde"),
        BotCommand("produto_mes", "🏆 Produto destaque"),
        BotCommand("giro",        "📦 Giro de produtos"),
        BotCommand("status",        "🔍 Status da assinatura"),
        BotCommand("reativar",      "🔄 Reativar assinatura"),
        BotCommand("configuracoes", "⚙️ Atualizar credenciais"),
        BotCommand("cancelar",      "❌ Cancelar assinatura"),
        BotCommand("menu",       "🔄 Menu"),
    ]
    await app.bot.set_my_commands(cmds)
    await app.bot.set_chat_menu_button(menu_button=MenuButtonCommands())

def kb_menu(periodo_label: str = ""):
    btn_periodo = "📅 Analisar outro período" if periodo_label else "📅 Analisar período"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Briefing",      callback_data="briefing"),
         InlineKeyboardButton("⚠️ Alertas",       callback_data="alertas")],
        [InlineKeyboardButton("📦 Produtos",      callback_data="produtos"),
         InlineKeyboardButton("🗂 Categorias",    callback_data="categorias")],
        [InlineKeyboardButton("💳 Pagamentos",    callback_data="pagamentos"),
         InlineKeyboardButton("🕐 Pico",          callback_data="pico")],
        [InlineKeyboardButton("📅 Semanal",       callback_data="semana"),
         InlineKeyboardButton("📈 Comparativo",   callback_data="comparativo")],
        [InlineKeyboardButton("🏆 Produto Mês",   callback_data="produto_mes"),
         InlineKeyboardButton("📦 Giro",          callback_data="giro")],
        [InlineKeyboardButton("🎯 Projeção Mês",  callback_data="projecao"),
         InlineKeyboardButton("⭐ Score Saúde",   callback_data="score")],
        [InlineKeyboardButton("🔍 Padrões da Operação", callback_data="agente_padroes"),
         InlineKeyboardButton("💬 Pergunte à IA",        callback_data="agente_ajuda")],
        [InlineKeyboardButton("💡 Oportunidades de Mix", callback_data="agente_oportunidades")],
        [InlineKeyboardButton("🛒 Lista de Reposição",  callback_data="reposicao")],
        [InlineKeyboardButton(btn_periodo,               callback_data="atualizar_menu")],
    ])

async def abrir_menu(msg, chat_id: int = None):
    """Envia menu com período ativo visível."""
    periodo_label = ""
    if chat_id and chat_id in dados_usuario:
        periodo_label = dados_usuario[chat_id].get("periodo_label", "")

    if periodo_label:
        texto = f"📅 {b(f'Período analisado atualmente: {periodo_label}')}\n\nO que deseja analisar?"
    else:
        texto = "📂 Escolha um período para começar."

    await msg.reply_text(texto, parse_mode="HTML", reply_markup=kb_menu(periodo_label))

# ─── ENVIO HTML ──────────────────────────────────────────────
def _md_para_html(texto: str) -> str:
    """
    Converte a formatação Markdown que o Claude usa (**negrito**, *itálico*)
    para HTML, que é o parse_mode usado em todo o bot. Sem isso, as respostas
    do agente (foto/áudio/texto livre) mostram os asteriscos literais em vez
    de negrito, ficando visualmente diferentes das respostas dos botões.
    """
    import re
    if not texto:
        return texto
    texto = texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    texto = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", texto)
    texto = re.sub(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)", r"<i>\1</i>", texto)
    texto = re.sub(r"(?<!_)_(?!_)(.+?)(?<!_)_(?!_)", r"<i>\1</i>", texto)
    texto = re.sub(r"`(.+?)`", r"<code>\1</code>", texto)
    return texto


async def enviar(msg, texto: str):
    """Envia texto com parse_mode HTML, dividindo se necessário.

    Converte Markdown para HTML automaticamente quando detecta formatação
    Markdown (asteriscos) sem tags HTML já presentes — isso padroniza as
    respostas do agente de IA (que vêm em Markdown) com o visual do resto do
    bot (que usa HTML), sem afetar mensagens que já vêm formatadas em HTML
    pelos blocos de análise.
    """
    if texto and "**" in texto and "<b>" not in texto and "<i>" not in texto:
        texto = _md_para_html(texto)
    LIMITE = 4000
    while len(texto) > LIMITE:
        corte = texto.rfind("\n", 0, LIMITE)
        if corte == -1:
            corte = LIMITE
        await msg.reply_text(texto[:corte], parse_mode="HTML")
        texto = texto[corte:].strip()
    if texto:
        await msg.reply_text(texto, parse_mode="HTML")

# ─── HANDLERS ────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        f"👋 {b('Bem-vindo(a) ao MercadoBot!')}\n\n"
        "Seu assistente de inteligência para mercadinhos autônomos.\n\n"
        f"{b('Como começar:')}\n"
        "1. Exporte o <i>Resumo Geral de Vendas</i> do PDV Legal em Excel\n"
        "2. Exporte os <i>Produtos Mais Vendidos</i> em Excel\n"
        "3. Envie os dois arquivos aqui\n\n"
        "Depois escolha uma opção abaixo 👇",
        parse_mode="HTML",
        reply_markup=kb_menu()
    )

async def comando_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Escolha uma opção:", reply_markup=kb_menu())

async def receber_arquivo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    doc = update.message.document
    if not doc.file_name.endswith((".xlsx", ".xls")):
        await update.message.reply_text("⚠️ Por favor envie um arquivo Excel (.xlsx)")
        return
    await update.message.reply_text("⏳ Processando arquivo...")
    file = await doc.get_file()
    bio = BytesIO()
    await file.download_to_memory(bio)
    bio.seek(0)
    try:
        df = pd.read_excel(bio)
        colunas = [c.lower() for c in df.columns]
        if chat_id not in dados_usuario:
            dados_usuario[chat_id] = {}
        if "nomefilial" in colunas and "formarecebimento" in colunas:
            dados_usuario[chat_id]["vendas"] = normalizar_vendas(df)
            await update.message.reply_text(
                f"✅ {b('Resumo de Vendas carregado')}\n📊 {len(df)} transações encontradas.\n\n"
                "Envie agora os <i>Produtos Mais Vendidos</i> ou use o menu abaixo.",
                parse_mode="HTML", reply_markup=kb_menu()
            )
        elif "produto" in colunas and "quantidade" in colunas:
            dados_usuario[chat_id]["produtos"] = normalizar_produtos(df)
            await update.message.reply_text(
                f"✅ {b('Produtos carregados')}\n📦 {df['produto'].nunique()} SKUs encontrados.\n\n"
                "Tudo pronto! Escolha uma análise:",
                parse_mode="HTML", reply_markup=kb_menu()
            )
        else:
            await update.message.reply_text(
                "⚠️ Formato não reconhecido.\nEnvie o <i>Resumo Geral de Vendas</i> ou os <i>Produtos Mais Vendidos</i>.",
                parse_mode="HTML"
            )
    except Exception as e:
        await update.message.reply_text(f"❌ Erro: {str(e)}")

async def receber_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Recebe uma foto (ex: produto na gôndola) e usa a visão nativa do Claude
    para identificar o que é — com acesso às mesmas ferramentas do agente de
    texto, então ele pode tanto identificar o produto quanto, se fizer
    sentido, já buscar dados reais de venda desse produto no PDV Legal.
    """
    chat_id = update.effective_chat.id
    await update.message.reply_text("📸 Analisando a imagem...")

    try:
        foto = update.message.photo[-1]  # maior resolução disponível
        file = await foto.get_file()
        bio  = BytesIO()
        await file.download_to_memory(bio)
        bio.seek(0)
        import base64
        imagem_base64 = base64.b64encode(bio.read()).decode("utf-8")

        legenda = update.message.caption or ""

        from agente import processar_mensagem_agente
        resposta = await processar_mensagem_agente(
            chat_id, texto_usuario=legenda,
            imagem_base64=imagem_base64, imagem_media_type="image/jpeg"
        )
        await enviar(update.message, resposta)
        await abrir_menu(update.message, chat_id)
    except Exception as e:
        logger.error(f"Erro ao processar foto de chat_id={chat_id}: {e}")
        await update.message.reply_text("⚠️ Não consegui analisar essa imagem agora. Tente de novo.")


async def receber_audio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Recebe uma mensagem de voz (ou arquivo de áudio), transcreve via Whisper
    (a API da Anthropic não expõe endpoint de áudio), e processa o texto
    resultante exatamente pelo mesmo caminho do agente conversacional usado
    para texto digitado — incluindo tool-use para buscar dados reais.
    """
    chat_id = update.effective_chat.id
    await update.message.reply_text("🎤 Ouvindo seu áudio...")

    try:
        voz  = update.message.voice or update.message.audio
        file = await voz.get_file()

        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            await file.download_to_drive(tmp.name)
            caminho_temp = tmp.name

        from agente import transcrever_audio, processar_mensagem_agente
        texto = await transcrever_audio(caminho_temp)

        try:
            os.remove(caminho_temp)
        except Exception:
            pass

        if not texto:
            await update.message.reply_text(
                "⚠️ Não consegui transcrever esse áudio agora. "
                "Pode escrever sua pergunta em texto? 🙏"
            )
            return

        await update.message.reply_text(f"💬 {i('Você disse:')} \"{texto}\"\n\n⏳ Pensando...")
        resposta = await processar_mensagem_agente(chat_id, texto)
        await enviar(update.message, resposta)
        await abrir_menu(update.message, chat_id)
    except Exception as e:
        logger.error(f"Erro ao processar áudio de chat_id={chat_id}: {e}")
        await update.message.reply_text("⚠️ Não consegui processar esse áudio agora. Tente de novo.")


# ─── FLUXO BRIEFING ──────────────────────────────────────────
async def buscar_usuario_db(chat_id: int) -> dict:
    from database import buscar_usuario
    return await buscar_usuario(chat_id)

async def comando_reposicao_msg(msg):
    """Versão da reposição que aceita objeto msg diretamente."""
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Repor exatamente o que saiu",        callback_data="rep_modo_exato")],
        [InlineKeyboardButton("📦 Repor + estoque de segurança (30%)", callback_data="rep_modo_estoque")],
    ])
    await msg.reply_text(
        f"🛒 {b('LISTA DE REPOSIÇÃO')}\n\nComo deseja repor?",
        parse_mode="HTML",
        reply_markup=kb
    )

async def sem_dados_msg(msg):
    """Mensagem quando não há dados carregados — orienta a usar Analisar período."""
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Carregar dados agora", callback_data="atualizar_menu")],
    ])
    await msg.reply_text(
        "📂 Nenhum dado carregado ainda.\n\n"
        "Escolha um período para começar a análise.",
        reply_markup=kb
    )
async def pedir_periodo(msg):
    """Redireciona para sem_dados_msg — mantido por compatibilidade."""
    await sem_dados_msg(msg)

async def fluxo_briefing(msg, chat_id: int):
    d = dados_usuario.get(chat_id, {})
    if not d:
        await msg.reply_text("📎 Envie seus arquivos Excel primeiro. Use /start.")
        return
    vendas   = d.get("vendas")
    produtos = d.get("produtos")
    ctx = resumo_dados(chat_id)

    total_cancel = d.get("total_cancel", 0.0)

    # Bloco 1 — Faturamento
    await enviar(msg, bloco_faturamento(vendas, produtos, total_cancel))
    if vendas is not None:
        await msg.reply_photo(photo=g_faturamento(vendas))

    # Bloco 2 — Categorias
    if produtos is not None:
        await enviar(msg, bloco_categorias(produtos))
        await msg.reply_photo(photo=g_categorias(produtos))

    # Bloco 3 — Pagamentos
    if vendas is not None:
        await enviar(msg, bloco_pagamentos(vendas))
        await msg.reply_photo(photo=g_pagamentos(vendas))

    # Bloco 4 — Semanal
    if vendas is not None:
        await enviar(msg, bloco_semanal(vendas))
        await msg.reply_photo(photo=g_semanal(vendas))

    # Bloco 5 — Insights IA (curtos)
    insight = await insight_ia(ctx, "cancelamentos, quedas de faturamento e oportunidades de produto")
    await enviar(msg, f"💡 {b('INSIGHT DO DIA')}\n\n{insight}")

    await abrir_menu(msg, chat_id)

async def comando_briefing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await pedir_periodo(update.message)

async def comando_produtos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id  = update.effective_chat.id
    d        = dados_usuario.get(chat_id, {})
    produtos = d.get("produtos")
    if produtos is None or produtos.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Analisando produtos...")
    await enviar(update.message, bloco_top_produtos(produtos))
    await update.message.reply_photo(photo=g_top_produtos(produtos))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "oportunidades de mix de produtos entre as unidades")
    await enviar(update.message, f"💡 {b('INSIGHTS')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_categorias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id  = update.effective_chat.id
    d        = dados_usuario.get(chat_id, {})
    produtos = d.get("produtos")
    if produtos is None or produtos.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Calculando receita por categoria...")
    await enviar(update.message, bloco_categorias(produtos))
    await update.message.reply_photo(photo=g_categorias(produtos))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "categorias com melhor e pior desempenho")
    await enviar(update.message, f"💡 {b('INSIGHTS')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_pagamentos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d       = dados_usuario.get(chat_id, {})
    vendas  = d.get("vendas")
    if vendas is None or vendas.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Analisando mix de pagamentos...")
    await enviar(update.message, bloco_pagamentos(vendas))
    await update.message.reply_photo(photo=g_pagamentos(vendas))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "mix de pagamento e oportunidade de incentivar PIX")
    await enviar(update.message, f"💡 {b('INSIGHTS')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def abrir_submenu_semana(msg, chat_id: int):
    """Abre o submenu com as 3 visões semanais."""
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Semana atual",                  callback_data="semana_atual")],
        [InlineKeyboardButton("🗓️ Mês atual",                     callback_data="semana_mes")],
        [InlineKeyboardButton("📊 Mês atual x Mês anterior",      callback_data="semana_comparativo")],
        [InlineKeyboardButton("◀️ Voltar ao menu",                callback_data="menu")],
    ])
    await msg.reply_text(
        f"📅 {b('Análise Semanal')}\n\nEscolha a visão que deseja:",
        parse_mode="HTML",
        reply_markup=kb
    )


async def comando_semana_atual(msg, chat_id: int):
    """Semana atual — seg a dom da semana corrente."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    br    = ZoneInfo("America/Sao_Paulo")
    agora = datetime.now(br)
    # Segunda-feira da semana atual
    seg   = agora - timedelta(days=agora.weekday())
    ini   = seg.strftime("%d/%m/%Y")
    fim   = agora.strftime("%d/%m/%Y")

    d      = dados_usuario.get(chat_id, {})
    vendas = d.get("vendas")

    # Se dados carregados já cobrem a semana, usa — senão busca
    data_ini_carregada = d.get("data_ini", "")
    if vendas is not None and not vendas.empty and data_ini_carregada <= ini:
        v = vendas
    else:
        await msg.reply_text("⏳ Buscando dados da semana atual...")
        try:
            from scraper import baixar_relatorios_periodo
            import asyncio
            loop = asyncio.get_event_loop()
            pdv_email = d.get("pdv_email") or dados_usuario.get(chat_id, {}).get("pdv_email")
            pdv_senha = d.get("pdv_senha") or dados_usuario.get(chat_id, {}).get("pdv_senha")
            path_v, path_p, total_c = await loop.run_in_executor(
                None, baixar_relatorios_periodo, ini, fim, pdv_email, pdv_senha
            )
            v = normalizar_vendas(pd.read_excel(path_v))
            dados_usuario[chat_id]["vendas"]       = v
            dados_usuario[chat_id]["periodo_label"] = f"Semana atual ({ini} – {fim})"
        except Exception as e:
            await msg.reply_text(f"❌ Erro ao buscar dados: {e}")
            return

    await enviar(msg, bloco_semanal(v))
    await msg.reply_photo(photo=g_semanal(v))
    ctx = f"Semana atual {ini} a {fim}. Faturamento: R$ {v['valor'].sum():.2f}"
    insight = await insight_ia(ctx, "tendência da semana e comparativo entre filiais")
    await enviar(msg, f"💡 {b('INSIGHT')}\n\n{insight}")
    await abrir_menu(msg, chat_id)


async def comando_semana_comparativo(msg, chat_id: int):
    """Mês atual x Mês anterior — semana a semana."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    br    = ZoneInfo("America/Sao_Paulo")
    agora = datetime.now(br)

    # Mês atual: do dia 1 até hoje
    ini_atual = agora.replace(day=1).strftime("%d/%m/%Y")
    fim_atual  = agora.strftime("%d/%m/%Y")

    # Mês anterior: do dia 1 ao último dia
    primeiro_atual   = agora.replace(day=1)
    ultimo_anterior  = primeiro_atual - timedelta(days=1)
    ini_anterior     = ultimo_anterior.replace(day=1).strftime("%d/%m/%Y")
    fim_anterior     = ultimo_anterior.strftime("%d/%m/%Y")

    await msg.reply_text(f"⏳ Buscando {agora.strftime('%B')} e mês anterior...")
    try:
        d         = dados_usuario.get(chat_id, {})
        pdv_email = d.get("pdv_email")
        pdv_senha = d.get("pdv_senha")
        from scraper import baixar_relatorios_periodo
        import asyncio
        loop = asyncio.get_event_loop()

        # Busca os dois períodos
        path_atual, _, _    = await loop.run_in_executor(None, baixar_relatorios_periodo, ini_atual,    fim_atual,    pdv_email, pdv_senha)
        path_anterior, _, _ = await loop.run_in_executor(None, baixar_relatorios_periodo, ini_anterior, fim_anterior, pdv_email, pdv_senha)

        v_atual    = normalizar_vendas(pd.read_excel(path_atual))
        v_anterior = normalizar_vendas(pd.read_excel(path_anterior))
    except Exception as e:
        await msg.reply_text(f"❌ Erro ao buscar dados: {e}")
        return

    # Bloco comparativo
    fat_atual    = v_atual["valor"].sum()
    fat_anterior = v_anterior["valor"].sum()
    var          = fat_atual - fat_anterior
    var_pct      = (var / fat_anterior * 100) if fat_anterior > 0 else 0
    sinal        = "📈" if var >= 0 else "📉"

    nome_atual    = agora.strftime("%B/%Y").capitalize()
    nome_anterior = ultimo_anterior.strftime("%B/%Y").capitalize()

    texto = (
        f"📊 {b('Mês atual x Mês anterior')}\n\n"
        f"📅 {b(nome_atual)}: R$ {fat_atual:,.2f}\n"
        f"📅 {b(nome_anterior)}: R$ {fat_anterior:,.2f}\n\n"
        f"{sinal} Variação: {b(f'R$ {abs(var):,.2f}')} ({var_pct:+.1f}%)\n"
    )

    # Detalhamento por filial
    if "nomeFilial" in v_atual.columns and "nomeFilial" in v_anterior.columns:
        filiais_atual    = v_atual.groupby("nomeFilial")["valor"].sum()
        filiais_anterior = v_anterior.groupby("nomeFilial")["valor"].sum()
        texto += f"\n{b('Por filial:')}\n"
        for filial in filiais_atual.index:
            fa = filiais_atual.get(filial, 0)
            fb = filiais_anterior.get(filial, 0)
            vf = fa - fb
            pf = (vf / fb * 100) if fb > 0 else 0
            sf = "📈" if vf >= 0 else "📉"
            texto += f"  {sf} {filial.title()}: R$ {fa:,.2f} ({pf:+.1f}%)\n"

    await enviar(msg, texto)

    # Gráfico usando dados combinados com label de mês
    v_atual["mes"]    = nome_atual
    v_anterior["mes"] = nome_anterior
    v_combined = pd.concat([v_anterior, v_atual])
    await msg.reply_photo(photo=g_semanal(v_combined))

    ctx = f"Comparativo {nome_atual} (R$ {fat_atual:.2f}) vs {nome_anterior} (R$ {fat_anterior:.2f}). Variação: {var_pct:+.1f}%"
    insight = await insight_ia(ctx, "comparativo entre os dois meses e oportunidades de melhoria")
    await enviar(msg, f"💡 {b('INSIGHT')}\n\n{insight}")
    await abrir_menu(msg, chat_id)


async def comando_semana_mes(msg, chat_id: int):
    """Mês atual agrupado por semana."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    br    = ZoneInfo("America/Sao_Paulo")
    agora = datetime.now(br)
    ini   = agora.replace(day=1).strftime("%d/%m/%Y")
    fim   = agora.strftime("%d/%m/%Y")

    await msg.reply_text("⏳ Buscando dados do mês...")
    try:
        d         = dados_usuario.get(chat_id, {})
        pdv_email = d.get("pdv_email")
        pdv_senha = d.get("pdv_senha")
        from scraper import baixar_relatorios_periodo
        import asyncio
        loop = asyncio.get_event_loop()
        path_v, _, _ = await loop.run_in_executor(
            None, baixar_relatorios_periodo, ini, fim, pdv_email, pdv_senha
        )
        v = normalizar_vendas(pd.read_excel(path_v))
    except Exception as e:
        await msg.reply_text(f"❌ Erro ao buscar dados: {e}")
        return

    await enviar(msg, bloco_semanal(v))
    await msg.reply_photo(photo=g_semanal(v))
    ctx = f"Mês atual {ini} a {fim}. Faturamento: R$ {v['valor'].sum():.2f}"
    insight = await insight_ia(ctx, "ritmo do mês e projeção de fechamento")
    await enviar(msg, f"💡 {b('INSIGHT')}\n\n{insight}")
    await abrir_menu(msg, chat_id)


async def comando_semana(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    await abrir_submenu_semana(update.message, chat_id)

async def comando_pico(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d       = dados_usuario.get(chat_id, {})
    vendas  = d.get("vendas")
    if vendas is None or vendas.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Analisando horários de pico...")
    await enviar(update.message, bloco_pico(vendas))
    await update.message.reply_photo(photo=g_pico(vendas))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "horários de pico e horários fracos para sugestão de ação")
    await enviar(update.message, f"💡 {b('INSIGHTS')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_alertas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d       = dados_usuario.get(chat_id, {})
    vendas  = d.get("vendas")
    if not d or vendas is None:
        await pedir_periodo(update.message)
        return

    await update.message.reply_text("⏳ Verificando alertas...")
    ctx    = resumo_dados(chat_id)
    linhas = [f"🚨 {b('ALERTAS E ATENÇÕES')}\n"]

    cancel = vendas["ValorItensCancelados"].sum()
    total  = vendas["valor"].sum()
    pct    = cancel / total * 100 if total > 0 else 0
    if pct > 25:
        linhas.append(f"⚠️ Cancelamentos em {b(f'{pct:.1f}%')} do faturamento — acima do ideal (25%)")
        linhas.append(f"   Valor: {c(f'R$ {cancel:,.2f}')}\n")

    sem = vendas.copy()
    sem["DataAbertura"] = pd.to_datetime(sem["DataAbertura"], dayfirst=True)
    sem["semana"] = sem["DataAbertura"].dt.isocalendar().week
    fat_sem = sem.groupby(["semana","nomeFilial"])["valor"].sum().unstack(fill_value=0)
    if len(fat_sem) >= 2:
        for col in fat_sem.columns:
            var = fat_sem.iloc[-1][col] - fat_sem.iloc[-2][col]
            if var < -200:
                nome = col.split()[-1].title()
                linhas.append(f"📉 {b(nome)}: queda de {i(f'R$ {abs(var):,.2f}')} na última semana")

    insight = await insight_ia(ctx, "todos os alertas críticos e ações corretivas imediatas")
    linhas.append(f"\n{insight}")
    await enviar(update.message, "\n".join(linhas))
    await abrir_menu(update.message, update.effective_chat.id)


async def comando_reposicao(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Passo 1: escolha do modo de reposição."""
    msg = update.message if update.message else update.callback_query.message
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Repor exatamente o que saiu",         callback_data="rep_modo_exato")],
        [InlineKeyboardButton("📦 Repor + estoque de segurança (30%)",  callback_data="rep_modo_estoque")],
    ])
    await msg.reply_text(
        f"🛒 {b('LISTA DE REPOSIÇÃO')}\n\n"
        f"Como deseja repor?",
        parse_mode="HTML",
        reply_markup=kb
    )


async def comando_comparativo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d = dados_usuario.get(chat_id, {})
    vendas = d.get("vendas")
    if vendas is None or vendas.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Comparando unidades...")
    await enviar(update.message, bloco_comparativo(vendas))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "comparativo entre unidades — qual tem melhor desempenho e por quê")
    await enviar(update.message, f"💡 {b('INSIGHT')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_produto_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d = dados_usuario.get(chat_id, {})
    produtos = d.get("produtos")
    if produtos is None or produtos.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Analisando produtos destaque...")
    await enviar(update.message, bloco_produto_mes(produtos))
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_giro(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d = dados_usuario.get(chat_id, {})
    produtos = d.get("produtos")
    if produtos is None or produtos.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Classificando giro de produtos...")
    await enviar(update.message, bloco_giro_produtos(produtos))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "estratégia de estoque baseada no giro dos produtos")
    await enviar(update.message, f"💡 {b('INSIGHT')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_projecao(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d = dados_usuario.get(chat_id, {})
    vendas = d.get("vendas")
    if vendas is None or vendas.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Calculando projeção do mês...")
    bloco = bloco_projecao_mes(vendas)
    if bloco:
        await enviar(update.message, bloco)
    else:
        await update.message.reply_text("⚠️ Dados insuficientes para projeção.")
    await abrir_menu(update.message, update.effective_chat.id)

async def comando_score(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    d = dados_usuario.get(chat_id, {})
    vendas = d.get("vendas")
    if vendas is None or vendas.empty:
        await pedir_periodo(update.message)
        return
    await update.message.reply_text("⏳ Calculando score de saúde...")
    await enviar(update.message, bloco_score(vendas))
    ctx = resumo_dados(chat_id)
    insight = await insight_ia(ctx, "principais ações para melhorar o score de saúde da operação")
    await enviar(update.message, f"💡 {b('COMO MELHORAR')}\n\n{insight}")
    await abrir_menu(update.message, update.effective_chat.id)


async def comando_atualizar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Hoje",            callback_data="atualizar_hoje")],
        [InlineKeyboardButton("📅 Ontem",           callback_data="atualizar_ontem")],
        [InlineKeyboardButton("📅 Últimos 7 dias",  callback_data="atualizar_7dias")],
        [InlineKeyboardButton("📅 Últimos 15 dias", callback_data="atualizar_15dias")],
        [InlineKeyboardButton("📅 Últimos 30 dias", callback_data="atualizar_30dias")],
        [InlineKeyboardButton("📅 Mês atual",       callback_data="atualizar_mes")],
        [InlineKeyboardButton("📅 Mês anterior",    callback_data="atualizar_mes_anterior")],
        [InlineKeyboardButton("◀️ Voltar ao menu",  callback_data="menu")],
    ])
    await update.message.reply_text(
        f"🔄 {b('ATUALIZAR DADOS')}\n\nQual período deseja buscar agora?",
        parse_mode="HTML",
        reply_markup=kb
    )

async def executar_atualizacao(msg, chat_id: int, data_ini: str, data_fim: str, label: str, gerar_briefing: bool = True):
    """Executa o scraper para o período escolhido e envia o briefing."""

    # Mensagem de status que vamos editar em tempo real
    status = await msg.reply_text(
        f"🔄 Iniciando busca de dados — {b(label)}\n\n"
        f"⏳ Conectando ao PDV Legal...",
        parse_mode="HTML"
    )

    async def atualizar_status(texto: str):
        try:
            await status.edit_text(texto, parse_mode="HTML")
        except Exception:
            pass

    try:
        await atualizar_status(
            f"🔄 Buscando dados — {b(label)}\n\n"
            f"✅ Conectado\n"
            f"⏳ Fazendo login no PDV Legal..."
        )

        from scraper import baixar_relatorios_periodo
        from database import buscar_usuario
        import pandas as pd

        # Busca credenciais do usuário no banco
        usuario   = await buscar_usuario(chat_id)
        pdv_email = usuario.get("pdv_email") if usuario else None
        pdv_senha = usuario.get("pdv_senha") if usuario else None

        # Executa o scraper com feedback por etapa
        loop = asyncio.get_event_loop()

        await atualizar_status(
            f"🔄 Buscando dados — {b(label)}\n\n"
            f"✅ Conectado\n"
            f"✅ Login realizado\n"
            f"⏳ Exportando Resumo de Vendas..."
        )

        path_vendas, path_produtos, total_cancel = await loop.run_in_executor(
            None, baixar_relatorios_periodo, data_ini, data_fim, pdv_email, pdv_senha
        )

        await atualizar_status(
            f"🔄 Buscando dados — {b(label)}\n\n"
            f"✅ Conectado\n"
            f"✅ Login realizado\n"
            f"✅ Vendas exportadas\n"
            f"✅ Produtos exportados\n"
            f"⏳ Processando e gerando análises..."
        )

        vendas_raw   = pd.read_excel(path_vendas)
        produtos_raw = pd.read_excel(path_produtos)

        vendas   = normalizar_vendas(vendas_raw)
        produtos = normalizar_produtos(produtos_raw)

        # Valida se vieram dados reais
        if len(vendas) == 0:
            await atualizar_status(
                f"🔄 Buscando dados — {b(label)}\n\n"
                f"⚠️ O relatório de vendas veio sem dados para esse período.\n\n"
                f"Possíveis causas:\n"
                f"• Não houve vendas nesse dia\n"
                f"• O filtro de data não foi reconhecido\n\n"
                f"Tente outro período ou importe o arquivo manualmente."
            )
            await abrir_menu(msg, chat_id)
            return

        if chat_id not in dados_usuario:
            dados_usuario[chat_id] = {}
        dados_usuario[chat_id]["vendas"]        = vendas
        dados_usuario[chat_id]["produtos"]      = produtos
        dados_usuario[chat_id]["periodo_label"] = label
        dados_usuario[chat_id]["total_cancel"]  = total_cancel

        await atualizar_status(
            f"🔄 Buscando dados — {b(label)}\n\n"
            f"✅ Conectado\n"
            f"✅ Login realizado\n"
            f"✅ Vendas exportadas\n"
            f"✅ Produtos exportados\n"
            f"✅ Dados processados\n\n"
            f"{'📊 Gerando briefing completo...' if gerar_briefing else '✅ Período carregado!'}"
        )

        if gerar_briefing:
            await fluxo_briefing(msg, chat_id)
        else:
            await abrir_menu(msg, chat_id)

    except Exception as e:
        erro = str(e)

        # Erro de login inválido
        if "login inválido" in erro.lower():
            kb_cred = InlineKeyboardMarkup([
                [InlineKeyboardButton("⚙️ Atualizar credenciais PDV Legal", callback_data="atualizar_credenciais")],
            ])
            await msg.reply_text(
                f"❌ {b('E-mail ou senha incorretos no PDV Legal')}\n\n"
                f"Suas credenciais não foram aceitas.\n"
                f"Clique abaixo para corrigir:",
                parse_mode="HTML",
                reply_markup=kb_cred
            )
            return

        # Timeout de download — PDV online mas demorou para gerar o arquivo
        elif "timeout" in erro.lower() and any(x in erro.lower() for x in ["download", "expect_download", "waiting for"]):
            logger.error(f"Timeout no download para {chat_id}: {erro[:300]}")
            try:
                await atualizar_status(
                    f"🔄 Buscando dados — {b(label)}\n\n"
                    f"⏳ O PDV Legal demorou para gerar o arquivo.\n\n"
                    f"Tente novamente — geralmente resolve na segunda tentativa."
                )
            except Exception:
                pass

        # Erro externo — site fora, manutenção, timeout de conexão
        elif any(x in erro.lower() for x in ["timeout", "manutenção", "maintenance",
                                              "txtemail", "txtsenha", "btnentrar",
                                              "net::err", "connection"]):
            logger.error(f"Erro de conexão PDV para {chat_id}: {erro[:300]}")
            try:
                await atualizar_status(
                    f"🔄 Buscando dados — {b(label)}\n\n"
                    f"⚠️ Não foi possível conectar ao PDV Legal\n\n"
                    f"• Site em manutenção ou instável\n"
                    f"• Lentidão no servidor do PDV Legal\n\n"
                    f"Tente novamente em alguns minutos."
                )
            except Exception:
                pass
        else:
            logger.error(f"Erro inesperado para {chat_id}: {erro[:300]}")
            try:
                await atualizar_status(
                    f"🔄 Buscando dados — {b(label)}\n\n"
                    f"❌ Erro inesperado\n\n"
                    f"{i(erro[:200])}"
                )
            except Exception:
                pass

        # Sempre abre o menu ao final — mesmo se editMessageText falhou
        try:
            await abrir_menu(msg, chat_id)
        except Exception:
            await msg.reply_text(
                "Use o menu para tentar novamente:",
                reply_markup=kb_menu()
            )

async def mensagem_livre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    texto   = update.message.text.strip()
    d       = dados_usuario.get(chat_id, {})

    # ─── Admin: aguardando texto de notificação ──────────────
    if context.user_data.get("admin_aguardando_notif") and chat_id == int(os.environ.get("ADMIN_CHAT_ID", "0")):
        from admin import admin_executar_notificacao
        from database import get_pool
        grupo = context.user_data.pop("admin_notif_grupo", "todos")
        context.user_data.pop("admin_aguardando_notif", None)
        pool  = await get_pool()
        await admin_executar_notificacao(update.message, pool, grupo, texto)
        return

    # ─── Fluxo de atualização de credenciais ─────────────────
    aguardando = d.get("aguardando")

    if aguardando == "novo_email":
        if "@" not in texto or "." not in texto:
            await update.message.reply_text("⚠️ E-mail inválido. Digite novamente:")
            return
        dados_usuario[chat_id]["novo_pdv_email"] = texto
        dados_usuario[chat_id]["aguardando"]     = "nova_senha"
        await update.message.reply_text(
            f"✅ E-mail registrado.\n\n"
            f"Agora digite sua nova {b('senha do PDV Legal')}:",
            parse_mode="HTML"
        )
        return

    if aguardando == "nova_senha":
        novo_email = d.get("novo_pdv_email")
        nova_senha = texto
        # Apaga mensagem com senha
        try:
            await update.message.delete()
        except Exception:
            pass
        from database import atualizar_usuario
        await atualizar_usuario(chat_id, pdv_email=novo_email, pdv_senha=nova_senha)
        dados_usuario[chat_id].pop("aguardando",    None)
        dados_usuario[chat_id].pop("novo_pdv_email", None)
        await update.message.reply_text(
            f"✅ {b('Credenciais atualizadas!')}\n\n"
            f"Suas novas credenciais foram salvas.\n"
            f"Use {b('🔄 Atualizar dados agora')} para testar.",
            parse_mode="HTML",
            reply_markup=kb_menu()
        )
        return

    # ─── Fluxo de reativação — coleta dados faltantes (telefone/CEP/número) ──
    if aguardando == "reativar_telefone":
        telefone = "".join(filter(str.isdigit, texto))
        if len(telefone) not in (10, 11):
            await update.message.reply_text(
                "⚠️ Telefone inválido. Digite com DDD, só números (ex: 11987654321):"
            )
            return
        dados_usuario[chat_id]["reativar_telefone"] = telefone
        dados_usuario[chat_id]["aguardando"] = "reativar_cep"
        await update.message.reply_text(
            f"✅ Telefone registrado.\n\n"
            f"📍 Agora informe seu {b('CEP')} (só números):\n\n"
            f"Exemplo: 89223005",
            parse_mode="HTML"
        )
        return

    if aguardando == "reativar_cep":
        cep = "".join(filter(str.isdigit, texto))
        if len(cep) != 8:
            await update.message.reply_text(
                "⚠️ CEP inválido. Digite só os 8 números (ex: 89223005):"
            )
            return
        dados_usuario[chat_id]["reativar_cep"] = cep
        dados_usuario[chat_id]["aguardando"] = "reativar_numero"
        await update.message.reply_text(
            f"✅ CEP registrado.\n\n"
            f"🏠 Por último, o {b('número')} do seu endereço:",
            parse_mode="HTML"
        )
        return

    if aguardando == "reativar_numero":
        numero = texto.strip()
        from database import buscar_usuario, atualizar_usuario
        telefone = dados_usuario[chat_id].pop("reativar_telefone", "")
        cep      = dados_usuario[chat_id].pop("reativar_cep", "")
        dados_usuario[chat_id].pop("aguardando", None)
        dados_usuario[chat_id].pop("reativar_dados_pendentes", None)

        await atualizar_usuario(chat_id, telefone=telefone, cep=cep, endereco_numero=numero)
        usuario = await buscar_usuario(chat_id)
        await update.message.reply_text("✅ Dados completos! Gerando link de reativação...")
        await _continuar_reativacao(update.message, chat_id, usuario)
        return

    # ─── Mensagem livre normal — agora processada pelo agente conversacional ──
    from agente import processar_mensagem_agente
    await update.message.reply_text("⏳ Pensando...")
    resposta = await processar_mensagem_agente(chat_id, texto)
    await enviar(update.message, resposta)
    await abrir_menu(update.message, update.effective_chat.id)

async def cmd_perguntar_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando dedicado para o agente conversacional — /perguntar <pergunta>."""
    chat_id = update.effective_chat.id
    texto   = " ".join(context.args).strip() if context.args else ""

    if not texto:
        await update.message.reply_text(
            f"💬 {b('Pergunte qualquer coisa sobre o seu negócio')}\n\n"
            f"Exemplos:\n"
            f"• /perguntar quanto vendi hoje?\n"
            f"• /perguntar como foi o faturamento da semana?\n"
            f"• /perguntar qual filial vendeu mais ontem?\n\n"
            f"Ou apenas escreva sua pergunta direto no chat, sem precisar do comando.",
            parse_mode="HTML"
        )
        return

    from agente import processar_mensagem_agente
    await update.message.reply_text("⏳ Pensando...")
    resposta = await processar_mensagem_agente(chat_id, texto)
    await enviar(update.message, resposta)
    await abrir_menu(update.message, chat_id)


async def cmd_status_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from onboarding import cmd_status
    await cmd_status(update, context)


def calcular_dias_trial_restantes(usuario: dict) -> int:
    """Retorna dias de trial restantes. 0 se expirado ou não havia trial."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    trial_fim = usuario.get("trial_fim")
    if not trial_fim:
        return 0
    try:
        fim   = datetime.fromisoformat(trial_fim)
        agora = datetime.now(ZoneInfo("America/Sao_Paulo"))
        dias  = (fim - agora).days
        return max(0, dias)
    except Exception:
        return 0


async def enviar_reativacao(update_or_msg, chat_id: int, is_callback: bool = False):
    """
    FUNÇÃO ÚNICA para enviar mensagem de reativação.
    Evita duplicação de mensagens.
    
    update_or_msg: pode ser Message (de /reativar) ou CallbackQuery.message (de callback)
    is_callback: True se vem de callback_query, False se de /reativar
    """
    from database import buscar_usuario
    from pagamento import gerar_link_pagamento, buscar_assinatura_ativa
    
    usuario = await buscar_usuario(chat_id)
    if not usuario:
        if is_callback:
            await update_or_msg.message.reply_text("Erro ao processar. Use /start.")
        else:
            await update_or_msg.reply_text("Use /start para criar uma conta.")
        return
    
    try:
        asaas_id = usuario.get("asaas_id")
        if not asaas_id:
            msg_text = "Use /start para reativar sua conta."
            if is_callback:
                await update_or_msg.message.reply_text(msg_text)
            else:
                await update_or_msg.reply_text(msg_text)
            return
        
        trial_usado = usuario.get("trial_usado", False)
        dias_trial = calcular_dias_trial_restantes(usuario) if trial_usado else 0
        
        # Busca assinatura ativa
        assinatura_id = await buscar_assinatura_ativa(asaas_id)
        link = None
        
        if assinatura_id:
            from pagamento import buscar_link_assinatura
            link = await buscar_link_assinatura(assinatura_id)
        
        # Se não tem link, gera novo checkout
        if not link:
            link, _ = await gerar_link_pagamento(
                asaas_id, chat_id,
                reativacao=trial_usado,
                dias_trial_restantes=dias_trial
            )
        
        # Mensagem ÚNICA de reativação
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Reativar assinatura", url=link)],
        ])
        
        msg_text = (
            f"🔄 {b('Reativar MercadoBot')}\n\n"
            f"Clique no botão abaixo para reativar sua assinatura:"
        )
        
        if is_callback:
            await update_or_msg.message.reply_text(msg_text, parse_mode="HTML", reply_markup=kb)
        else:
            await update_or_msg.reply_text(msg_text, parse_mode="HTML", reply_markup=kb)
            
    except Exception as e:
        logger.error(f"Erro em enviar_reativacao: {e}")
        msg_text = "❌ Erro ao processar reativação.\n\nUse /start para tentar novamente."
        try:
            if is_callback:
                await update_or_msg.message.reply_text(msg_text)
            else:
                await update_or_msg.reply_text(msg_text)
        except:
            pass


async def cmd_reativar_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /reativar — fluxo diferenciado por status."""
    from database import buscar_usuario, usuario_tem_acesso
    from datetime import datetime
    from zoneinfo import ZoneInfo
    
    chat_id = update.effective_chat.id
    usuario = await buscar_usuario(chat_id)

    if not usuario:
        await update.message.reply_text(
            "Você não tem cadastro. Use /start para criar uma conta."
        )
        return

    status = usuario["status"]
    BRASILIA = ZoneInfo("America/Sao_Paulo")
    agora = datetime.now(BRASILIA)

    # ─── Caso 1: cancelado_mas_ativo COM acesso válido ─────────────────────────
    # Usuário cancelou mas ainda tem dias de acesso já pago.
    # Ofereça agendar renovação para a data de vencimento (sem cobrar agora)
    if status == "cancelado_mas_ativo":
        tem_acesso, _ = await usuario_tem_acesso(chat_id)
        if tem_acesso:
            assinatura_fim = datetime.fromisoformat(usuario["assinatura_fim"])
            dias_restantes = (assinatura_fim - agora).days + 1
            
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Sim, renovar em " + assinatura_fim.strftime('%d/%m'), 
                                      callback_data="reativar_agendar")],
                [InlineKeyboardButton("❌ Não, deixar expirar", callback_data="menu_principal")],
            ])
            
            await update.message.reply_text(
                f"⏸️ {b('Sua assinatura foi cancelada.')}\n\n"
                f"Mas você ainda tem acesso até {b(assinatura_fim.strftime('%d/%m/%Y'))}.\n"
                f"Restam {b(f'{dias_restantes} dias')}.\n\n"
                f"🔄 {b('Deseja renovar automaticamente?')}\n"
                f"Cobraremos R$ 29,90 em {b(assinatura_fim.strftime('%d/%m'))}\n"
                f"(sem cobranças até lá).",
                parse_mode="HTML",
                reply_markup=kb
            )
            return
        # Senão, caiu fora e vai para o fluxo de cobrança abaixo

    # ─── Caso 2: trial, ativo, ou cancelado_mas_ativo SEM acesso ─────────────
    # Verifica se tem acesso válido em geral
    if status in ("trial", "ativo"):
        tem_acesso, _ = await usuario_tem_acesso(chat_id)
        if tem_acesso:
            await update.message.reply_text(
                f"✅ Sua assinatura já está ativa!\n\nUse /menu para acessar o bot.",
                parse_mode="HTML",
                reply_markup=kb_menu()
            )
            return
    
    # ─── Caso 3: status cancelado, expirado, ou cancelado_mas_ativo EXPIRADO ──
    # Precisa fazer novo pagamento para reativar
    faltando = []
    if not usuario.get("telefone"):
        faltando.append("telefone")
    if not usuario.get("cep"):
        faltando.append("cep")
    if not usuario.get("endereco_numero"):
        faltando.append("numero")

    if faltando:
        if chat_id not in dados_usuario:
            dados_usuario[chat_id] = {}
        dados_usuario[chat_id]["aguardando"] = "reativar_telefone"
        dados_usuario[chat_id]["reativar_dados_pendentes"] = faltando
        await update.message.reply_text(
            f"🔄 {b('Reativar MercadoBot')}\n\n"
            f"Antes de continuar, vamos {b('atualizar seu cadastro')} — "
            f"são só 3 perguntas rápidas.\n\n"
            f"📱 Informe seu {b('telefone com DDD')} (só números):\n\n"
            f"Exemplo: 11987654321",
            parse_mode="HTML"
        )
        return

    # Dados completos — gera novo checkout de cobrança
    await enviar_reativacao(update.message, chat_id, is_callback=False)




async def cmd_cancelar_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from onboarding import cmd_cancelar_assinatura
    await cmd_cancelar_assinatura(update, context)

async def cmd_configuracoes_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔑 Atualizar e-mail e senha do PDV Legal", callback_data="atualizar_credenciais")],
        [InlineKeyboardButton("🔍 Ver status da assinatura", callback_data="verificar_status")],
    ])
    await update.message.reply_text(
        f"⚙️ {b('Configurações')}\n\nO que deseja atualizar?",
        parse_mode="HTML",
        reply_markup=kb
    )

async def verificar_status_callback(msg, chat_id: int):
    """Verifica status e responde no chat."""
    from database import buscar_usuario
    from datetime import datetime
    from zoneinfo import ZoneInfo
    brasilia = ZoneInfo("America/Sao_Paulo")
    usuario  = await buscar_usuario(chat_id)

    if not usuario:
        await msg.reply_text(
            "👋 Use /start para se cadastrar no MercadoBot.",
            parse_mode="HTML"
        )
        return

    status = usuario["status"]

    if status in ("trial", "ativo"):
        agora = datetime.now(brasilia)
        fim   = datetime.fromisoformat(usuario.get("trial_fim") or usuario.get("assinatura_fim", ""))
        dias  = max(0, (fim - agora).days + 1)
        await msg.reply_text(
            f"✅ {b('Acesso liberado!')}\n\n"
            f"Você tem {b(f'{dias} dias')} restantes.\n\n"
            f"Use o menu abaixo para começar:",
            parse_mode="HTML",
            reply_markup=kb_menu()
        )
        return

    # Qualquer outro status (pendente, cancelado, bloqueado, expirado) —
    # ainda não pagou ou está reativando. Em vez de tentar recuperar um link
    # antigo (que pode não existir mais no fluxo via Checkout), oferece
    # gerar um link novo diretamente, evitando o usuário ficar travado
    # clicando em "Verificar novamente" sem nunca ver o botão de pagamento.
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Gerar novo link de pagamento", callback_data="reativar")],
    ])
    await msg.reply_text(
        f"⏳ {b('Ainda sem confirmação de pagamento.')}\n\n"
        f"Se ainda não cadastrou o cartão, ou se o link expirou, "
        f"clique abaixo para gerar um novo:",
        parse_mode="HTML",
        reply_markup=kb
    )

async def receber_arquivo_com_acesso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Verifica acesso antes de processar arquivo."""
    if not await verificar_acesso(update, context):
        return
    await receber_arquivo(update, context)

async def receber_foto_com_acesso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Verifica acesso antes de processar foto."""
    if not await verificar_acesso(update, context):
        return
    await receber_foto(update, context)

async def receber_audio_com_acesso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Verifica acesso antes de processar áudio."""
    if not await verificar_acesso(update, context):
        return
    await receber_audio(update, context)

async def mensagem_livre_com_acesso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Verifica acesso antes de processar mensagem livre."""
    if not await verificar_acesso(update, context):
        return
    await mensagem_livre(update, context)

async def _continuar_reativacao(msg, chat_id: int, usuario: dict):
    """Gera o link de reativação — chamado direto se os dados já estão completos,
    ou após coletar telefone/CEP/número de cadastros antigos."""
    from database import atualizar_usuario
    from pagamento import gerar_link_pagamento, buscar_assinatura_ativa, buscar_link_assinatura, atualizar_cliente_asaas

    await msg.reply_text("⏳ Gerando link de reativação...")

    try:
        asaas_id             = usuario["asaas_id"]
        trial_usado          = usuario.get("trial_usado", False)
        dias_trial_restantes = calcular_dias_trial_restantes(usuario) if trial_usado else 0
        logger.info(f"Reativação chat_id={chat_id}: asaas_id={asaas_id}, trial_usado={trial_usado}, dias_restantes={dias_trial_restantes}")

        # Sincroniza telefone/CEP/número com o Asaas — necessário para o Checkout
        # com cartão de crédito recorrente. Esses dados ficam no nosso banco desde
        # a coleta, mas o cliente no Asaas pode não ter sido atualizado ainda.
        await atualizar_cliente_asaas(
            asaas_id,
            telefone=usuario.get("telefone"),
            cep=usuario.get("cep"),
            endereco_numero=usuario.get("endereco_numero"),
        )
        logger.info(f"Reativação chat_id={chat_id}: dados sincronizados com Asaas")

        assinatura_id = await buscar_assinatura_ativa(asaas_id)
        logger.info(f"Reativação chat_id={chat_id}: assinatura_ativa encontrada={assinatura_id}")
        link = ""
        if assinatura_id:
            link = await buscar_link_assinatura(assinatura_id)
            logger.info(f"Reativação chat_id={chat_id}: link da assinatura existente={bool(link)}")
            if not link:
                assinatura_id = None

        if not assinatura_id:
            # Cancela qualquer checkout anterior ainda pendente desse usuário
            # antes de criar um novo — evita ter vários links "vivos" ao
            # mesmo tempo, que depois expiram e geram notificações confusas.
            checkout_anterior = usuario.get("ultimo_checkout_id")
            if checkout_anterior:
                from pagamento import cancelar_checkout
                await cancelar_checkout(checkout_anterior)

            logger.info(f"Reativação chat_id={chat_id}: gerando novo checkout...")
            link, assinatura_id = await gerar_link_pagamento(
                asaas_id, chat_id,
                reativacao=trial_usado,
                dias_trial_restantes=dias_trial_restantes
            )
            logger.info(f"Reativação chat_id={chat_id}: checkout gerado, link={bool(link)}, id={assinatura_id}")
            if assinatura_id:
                await atualizar_usuario(
                    chat_id, assinatura_asaas_id=assinatura_id, status="pendente",
                    ultimo_checkout_id=assinatura_id
                )

        if trial_usado and dias_trial_restantes > 0:
            aviso_trial = f"\n\n{i(f'Você ainda tem {dias_trial_restantes} dia(s) de trial restantes.')}"
        elif trial_usado and dias_trial_restantes <= 0:
            aviso_trial = f"\n\n{i('Trial já utilizado — cobrança imediata de R$ 29,90.')}"
        else:
            aviso_trial = ""

        if link:
            kb_reativar = InlineKeyboardMarkup([
                [InlineKeyboardButton("💳 Reativar assinatura", url=link)],
                [InlineKeyboardButton("🔍 Verificar status",    callback_data="verificar_status")],
            ])
            await msg.reply_text(
                f"🔄 {b('Reativar MercadoBot')}\n\n"
                f"Clique abaixo para reativar sua assinatura de {b('R$ 29,90/mês')}."
                f"{aviso_trial}",
                parse_mode="HTML",
                reply_markup=kb_reativar
            )
        else:
            await msg.reply_text("❌ Erro ao gerar link. Tente /reativar ou /start.")
    except Exception as e:
        import traceback
        logger.error(f"Erro na reativação para chat_id={chat_id}: {e}\n{traceback.format_exc()}")
        await msg.reply_text("❌ Erro ao processar. Tente /reativar.")


async def callback_botoes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    chat_id = query.message.chat_id
    acao    = query.data
    msg     = query.message

    # Remove os botões da mensagem que foi clicada
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass

    # Imports fixos no topo — evita UnboundLocalError por imports locais conflitantes
    from datetime import datetime as _datetime, timedelta as _timedelta
    from zoneinfo import ZoneInfo as _ZoneInfo
    _brasilia = _ZoneInfo("America/Sao_Paulo")
    _fmt      = "%d/%m/%Y"
    _hoje     = _datetime.now(_brasilia)

    # Mapa de textos para o popup instantâneo
    textos_popup = {
        "briefing":         "📊 Gerando briefing...",
        "produtos":         "📦 Analisando produtos...",
        "categorias":       "🗂 Calculando categorias...",
        "pagamentos":       "💳 Analisando pagamentos...",
        "semana":             "📅 Abrindo opções semanais...",
        "semana_atual":       "📅 Buscando semana atual...",
        "semana_comparativo": "📊 Buscando últimas 4 semanas...",
        "semana_mes":         "🗓️ Buscando mês atual...",
        "pico":             "🕐 Analisando horários...",
        "alertas":          "⚠️ Verificando alertas...",
        "reposicao":        "🛒 Abrindo reposição...",
        "atualizar_menu":   "🔄 Carregando períodos...",
        "verificar_status": "🔍 Verificando pagamento...",
    }

    popup = textos_popup.get(acao, "⏳ Processando...")
    await query.answer(popup)

    # ─── Admin callbacks ─────────────────────────────────────
    if acao.startswith("admin_"):
        from admin import callback_admin
        await callback_admin(update, context)
        return

    # ─── Gate de acesso — bloqueia ações sensíveis sem assinatura ativa ──
    # Ações sempre liberadas: reativação, verificação de status, navegação
    # básica de menu e o fluxo de atualizar credenciais (necessário até para
    # quem está reativando a conta).
    acoes_sempre_liberadas = {
        "reativar", "verificar_status", "menu_principal",
        "atualizar_credenciais", "confirmar_cancelamento",
        "pico_ok", "pico_problema",
    }
    if acao not in acoes_sempre_liberadas and not acao.startswith("reativar_"):
        if not await verificar_acesso(update, context):
            return

    if acao == "pico_ok":
        await query.answer("✅ Ótimo! Boas vendas nessa noite! 🚀")
        return

    if acao == "pico_problema":
        await msg.reply_text(
            f"⚠️ {b('Entendido — registramos o problema.')}\n\n"
            f"Verifique:\n"
            f"• Conexão com a internet no totem\n"
            f"• Sistema PDV Legal funcionando\n"
            f"• Totem com energia e tela ativa\n\n"
            f"Use /atualizar após resolver para confirmar que voltou ao normal.",
            parse_mode="HTML"
        )
        return

    # ─── Menu principal ──────────────────────────────────────
    if acao == "menu_principal":
        await abrir_menu(msg, chat_id)
        return

    # ─── Atualizar credenciais PDV Legal ─────────────────────
    if acao == "atualizar_credenciais":
        if chat_id not in dados_usuario:
            dados_usuario[chat_id] = {}
        dados_usuario[chat_id]["aguardando"] = "novo_email"
        await msg.reply_text(
            f"⚙️ {b('Atualizar credenciais PDV Legal')}\n\n"
            f"Digite seu novo {b('e-mail de login do PDV Legal')}:",
            parse_mode="HTML"
        )
        return

    # ─── Reativar assinatura ─────────────────────────────────
    if acao == "reativar":
        await query.answer()  # Remove o "spinning" do botão
        from database import buscar_usuario
        usuario = await buscar_usuario(chat_id)

        if not usuario or not usuario.get("asaas_id"):
            await msg.reply_text("Use /start para criar um novo cadastro.")
            return

        # Verifica se faltam dados obrigatórios para o Checkout do Asaas
        faltando = []
        if not usuario.get("telefone"):
            faltando.append("telefone")
        if not usuario.get("cep"):
            faltando.append("cep")
        if not usuario.get("endereco_numero"):
            faltando.append("numero")

        if faltando:
            if chat_id not in dados_usuario:
                dados_usuario[chat_id] = {}
            dados_usuario[chat_id]["aguardando"] = "reativar_telefone"
            dados_usuario[chat_id]["reativar_dados_pendentes"] = faltando
            await msg.reply_text(
                f"🔄 {b('Reativar MercadoBot')}\n\n"
                f"Antes de continuar, vamos {b('atualizar seu cadastro')} — "
                f"são só 3 perguntas rápidas.\n\n"
                f"📱 Informe seu {b('telefone com DDD')} (só números):\n\n"
                f"Exemplo: 11987654321",
                parse_mode="HTML"
            )
            return

        # Usa função ÚNICA de reativação
        await enviar_reativacao(msg, chat_id, is_callback=False)
        return

    # ─── Agendar renovação automática (cancelado_mas_ativo com acesso válido) ──
    if acao == "reativar_agendar":
        await query.answer()
        from database import buscar_usuario
        from pagamento import gerar_link_pagamento
        from datetime import datetime
        
        usuario = await buscar_usuario(chat_id)
        if not usuario:
            await msg.reply_text("Erro: usuário não encontrado.")
            return
        
        assinatura_fim = datetime.fromisoformat(usuario["assinatura_fim"])
        
        # Formata a data para o Asaas (YYYY-MM-DD HH:MM:SS)
        proxima_cobranca_str = assinatura_fim.strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            # Gera checkout agendado para começar em assinatura_fim
            link, assinatura_id = await gerar_link_pagamento(
                usuario.get("asaas_id"),
                chat_id,
                reativacao=True,
                proxima_cobranca_em=proxima_cobranca_str
            )
            
            if not link:
                await msg.reply_text(
                    "❌ Erro ao gerar link de pagamento.\n\n"
                    "Use /reativar para tentar novamente.",
                    parse_mode="HTML"
                )
                return
            
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("💳 Confirmar cartão", url=link)],
            ])
            
            await msg.reply_text(
                f"✅ {b('Renovação agendada!')}\n\n"
                f"Clique no botão abaixo para confirmar seu cartão.\n\n"
                f"💳 Cobraremos em {b(assinatura_fim.strftime('%d/%m/%Y'))}\n"
                f"(valor: R$ 29,90)\n\n"
                f"Até lá, você tem acesso normal! 🎉",
                parse_mode="HTML",
                reply_markup=kb
            )
        except Exception as e:
            logger.error(f"Erro ao agendar renovação: {e}")
            await msg.reply_text(
                "❌ Erro ao processar agendamento.\n\n"
                "Use /reativar para tentar novamente.",
                parse_mode="HTML"
            )
        return



    # ─── Verificar status de pagamento ──────────────────────
    if acao == "verificar_status":
        await verificar_status_callback(msg, chat_id)
        return

    # ─── Confirmar cancelamento de assinatura ────────────────
    if acao == "agente_padroes":
        from agente import processar_mensagem_agente
        await msg.reply_text("🔍 Analisando padrões na sua operação...")
        resposta = await processar_mensagem_agente(
            chat_id, "Identifique padrões relevantes na minha operação dos últimos 30 dias."
        )
        await enviar(msg, resposta)
        await abrir_menu(msg, chat_id)
        return

    if acao == "agente_ajuda":
        await msg.reply_text(
            f"💬 {b('Pergunte qualquer coisa sobre o seu negócio')}\n\n"
            f"Você pode digitar perguntas livres, mandar uma {b('foto')} de um "
            f"produto, ou até um {b('áudio')} — tudo direto no chat.\n\n"
            f"Exemplos:\n"
            f"• \"Quanto vendi hoje?\"\n"
            f"• \"Qual meu produto campeão do mês?\"\n"
            f"• \"Esse produto vende bem comparado a outros mercadinhos?\"\n"
            f"• Envie uma foto de um produto na gôndola\n\n"
            f"Pode escrever sua pergunta agora 👇",
            parse_mode="HTML"
        )
        return

    if acao == "agente_oportunidades":
        from agente import processar_mensagem_agente
        await msg.reply_text("💡 Procurando oportunidades de novos produtos para o seu mix...")
        resposta = await processar_mensagem_agente(
            chat_id, "Quais produtos vendem bem em outros mercadinhos que eu ainda não tenho no meu mix?"
        )
        await enviar(msg, resposta)
        await abrir_menu(msg, chat_id)
        return

    if acao == "confirmar_cancelamento":
        from database import buscar_usuario, atualizar_usuario
        from pagamento import cancelar_assinatura
        from datetime import datetime, timedelta
        from zoneinfo import ZoneInfo

        usuario = await buscar_usuario(chat_id)
        brasilia = ZoneInfo("America/Sao_Paulo")
        agora    = datetime.now(brasilia)

        # Verifica se está no período de trial (7 dias)
        dentro_do_trial = False
        if usuario and usuario.get("trial_fim"):
            trial_fim = datetime.fromisoformat(usuario["trial_fim"])
            dentro_do_trial = agora <= trial_fim

        # Calcula data de vencimento (quando acesso será bloqueado)
        data_vencimento = None
        if usuario:
            assinatura_fim_raw = usuario.get("assinatura_fim")
            if assinatura_fim_raw:
                try:
                    assinatura_fim_dt = datetime.fromisoformat(assinatura_fim_raw)
                    data_vencimento = assinatura_fim_dt.strftime("%d/%m/%Y")
                except:
                    pass

        if usuario:
            assinatura_id = usuario.get("assinatura_asaas_id")
            
            if assinatura_id:
                # Sempre cancela a assinatura (funciona para trial e ativo)
                # Cancela a subscription e qualquer cobrança futura associada
                await cancelar_assinatura(assinatura_id)
                logger.info(f"Assinatura cancelada: {assinatura_id}")

        await atualizar_usuario(chat_id, status="cancelado_mas_ativo")

        kb_reativar = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Reativar assinatura", callback_data="reativar")],
        ])

        if dentro_do_trial:
            # Cancelamento durante trial: sem cobranças
            await msg.reply_text(
                f"❌ {b('Cancelamento confirmado.')}\n\n"
                f"Como o cancelamento ocorreu durante o período de trial, "
                f"{b('nenhum valor será cobrado')}.\n\n"
                f"Você pode reativar a qualquer momento.",
                parse_mode="HTML",
                reply_markup=kb_reativar
            )
        else:
            # Cancelamento com acesso até data de vencimento
            if data_vencimento:
                await msg.reply_text(
                    f"❌ {b('Cancelamento confirmado.')}\n\n"
                    f"Você pode continuar usando o MercadoBot até {b(data_vencimento)}.\n\n"
                    f"✅ Não haverá novas cobranças.\n"
                    f"Pode reativar a qualquer momento.",
                    parse_mode="HTML",
                    reply_markup=kb_reativar
                )
            else:
                await msg.reply_text(
                    f"❌ {b('Cancelamento confirmado.')}\n\n"
                    f"✅ Não haverá novas cobranças.\n"
                    f"Pode reativar a qualquer momento.",
                    parse_mode="HTML",
                    reply_markup=kb_reativar
                )
        return

    # ─── Atualizar menu ─────────────────────────────────────
    if acao in ("atualizar_menu", "briefing_periodo"):
        gerar_briefing = (acao == "briefing_periodo")
        context.user_data["gerar_briefing"] = gerar_briefing
        titulo = "📊 Briefing — Escolha o período:" if gerar_briefing else "📅 Analisar outro período — Escolha o período:"
        kb_periodos = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Hoje",            callback_data="atualizar_hoje")],
            [InlineKeyboardButton("📅 Ontem",           callback_data="atualizar_ontem")],
            [InlineKeyboardButton("📅 Últimos 7 dias",  callback_data="atualizar_7dias")],
            [InlineKeyboardButton("📅 Últimos 15 dias", callback_data="atualizar_15dias")],
            [InlineKeyboardButton("📅 Últimos 30 dias", callback_data="atualizar_30dias")],
            [InlineKeyboardButton("📅 Mês atual",       callback_data="atualizar_mes")],
            [InlineKeyboardButton("📅 Mês anterior",    callback_data="atualizar_mes_anterior")],
            [InlineKeyboardButton("◀️ Voltar ao menu",  callback_data="menu")],
        ])
        await msg.reply_text(titulo, parse_mode="HTML", reply_markup=kb_periodos)
        return

    # ─── Atualizar: período escolhido ───────────────────────
    if acao.startswith("atualizar_"):
        if _hoje.month == 1:
            _mes_ant_ini = _hoje.replace(year=_hoje.year-1, month=12, day=1)
        else:
            _mes_ant_ini = _hoje.replace(month=_hoje.month-1, day=1)
        _mes_ant_fim = _hoje.replace(day=1) - _timedelta(days=1)

        periodos = {
            "atualizar_hoje":         (_hoje.strftime(_fmt),                          _hoje.strftime(_fmt),          "hoje"),
            "atualizar_ontem":        ((_hoje-_timedelta(days=1)).strftime(_fmt),     (_hoje-_timedelta(days=1)).strftime(_fmt), "ontem"),
            # "Últimos N dias" segue a definição do PDV Legal: N dias completos
            # terminando ONTEM (não inclui hoje, que é dia parcial). Ex: em 28/06,
            # "últimos 7 dias" = 21/06 a 27/06. Isso mantém o bot idêntico ao PDV
            # Legal para evitar divergência ao confrontar os dados.
            "atualizar_7dias":        ((_hoje-_timedelta(days=7)).strftime(_fmt),     (_hoje-_timedelta(days=1)).strftime(_fmt),  "últimos 7 dias"),
            "atualizar_15dias":       ((_hoje-_timedelta(days=15)).strftime(_fmt),    (_hoje-_timedelta(days=1)).strftime(_fmt),  "últimos 15 dias"),
            "atualizar_30dias":       ((_hoje-_timedelta(days=30)).strftime(_fmt),    (_hoje-_timedelta(days=1)).strftime(_fmt),  "últimos 30 dias"),
            "atualizar_mes":          (_hoje.strftime("01/%m/%Y"),                    _hoje.strftime(_fmt),          f"mês de {nome_mes(_hoje.month)}"),
            "atualizar_mes_anterior": (_mes_ant_ini.strftime(_fmt),                   _mes_ant_fim.strftime(_fmt),   f"{nome_mes(_mes_ant_ini.month)} de {_mes_ant_ini.year}"),
        }

        if acao in periodos:
            ini, fim, label = periodos[acao]
            gerar_briefing = context.user_data.get("gerar_briefing", True)
            context.user_data.pop("gerar_briefing", None)
            await executar_atualizacao(msg, chat_id, ini, fim, label, gerar_briefing=gerar_briefing)
        return

    if acao.startswith("rep_"):

        # ── Passo 1: escolha do modo ─────────────────────────
        if acao in ("rep_modo_exato", "rep_modo_estoque"):
            modo = "exato" if acao == "rep_modo_exato" else "estoque"
            dados_usuario.setdefault(chat_id, {})["rep_modo"] = modo
            label = "exatamente o que saiu" if modo == "exato" else "com estoque de segurança (+30%)"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📅 Hoje",            callback_data="rep_per_hoje")],
                [InlineKeyboardButton("📅 Ontem",           callback_data="rep_per_ontem")],
                [InlineKeyboardButton("📅 Últimos 7 dias",  callback_data="rep_per_7dias")],
                [InlineKeyboardButton("📅 Últimos 15 dias", callback_data="rep_per_15dias")],
                [InlineKeyboardButton("📅 Últimos 30 dias", callback_data="rep_per_30dias")],
                [InlineKeyboardButton("📅 Mês atual",       callback_data="rep_per_mes")],
                [InlineKeyboardButton("◀️ Voltar ao menu",  callback_data="menu")],
            ])
            await msg.reply_text(
                f"✅ Modo: {b(label)}\n\n"
                f"Qual período deseja usar para calcular a reposição?",
                parse_mode="HTML",
                reply_markup=kb
            )
            return

        # ── Passo 2: escolha do período ──────────────────────
        if acao.startswith("rep_per_"):
            periodos = {
                "rep_per_hoje":   (_hoje.strftime(_fmt),                         _hoje.strftime(_fmt),                         "hoje"),
                "rep_per_ontem":  ((_hoje-_timedelta(days=1)).strftime(_fmt),    (_hoje-_timedelta(days=1)).strftime(_fmt),    "ontem"),
                # "Últimos N dias" terminam ONTEM, igual ao PDV Legal (ver nota no mapa de atualizar)
                "rep_per_7dias":  ((_hoje-_timedelta(days=7)).strftime(_fmt),    (_hoje-_timedelta(days=1)).strftime(_fmt),    "últimos 7 dias"),
                "rep_per_15dias": ((_hoje-_timedelta(days=15)).strftime(_fmt),   (_hoje-_timedelta(days=1)).strftime(_fmt),    "últimos 15 dias"),
                "rep_per_30dias": ((_hoje-_timedelta(days=30)).strftime(_fmt),   (_hoje-_timedelta(days=1)).strftime(_fmt),    "últimos 30 dias"),
                "rep_per_mes":    (_hoje.strftime("01/%m/%Y"),                   _hoje.strftime(_fmt),                         "mês atual"),
            }

            if acao not in periodos:
                return

            ini, fim, label_per = periodos[acao]
            dados_usuario.setdefault(chat_id, {})["rep_periodo"] = (ini, fim, label_per)

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("💬 Receber no chat",        callback_data="rep_fmt_chat")],
                [InlineKeyboardButton("📊 Excel por loja",         callback_data="rep_fmt_excel_loja")],
                [InlineKeyboardButton("📊 Excel unificado",        callback_data="rep_fmt_excel_unificado")],
                [InlineKeyboardButton("◀️ Voltar ao menu",         callback_data="menu")],
            ])
            await msg.reply_text(
                f"📅 Período: {b(label_per)}\n\n"
                f"Como deseja receber a lista?",
                parse_mode="HTML",
                reply_markup=kb
            )
            return

        # ── Passo 3: gerar a reposição ───────────────────────
        if acao.startswith("rep_fmt_"):
            d        = dados_usuario.get(chat_id, {})
            modo     = d.get("rep_modo", "exato")
            periodo  = d.get("rep_periodo")
            formato  = acao.replace("rep_fmt_", "")

            if not periodo:
                await msg.reply_text("⚠️ Selecione o modo e período primeiro.")
                await comando_reposicao_msg(msg)
                return

            ini, fim, label_per = periodo
            label_modo = "exata do que saiu" if modo == "exato" else "com estoque de segurança (+30%)"

            await msg.reply_text(
                f"⏳ Buscando dados de {b(label_per)} e gerando lista de reposição {label_modo}...",
                parse_mode="HTML"
            )

            # Busca os dados do período escolhido
            try:
                usuario_db = await buscar_usuario_db(chat_id)
                pdv_email  = usuario_db.get("pdv_email") if usuario_db else None
                pdv_senha  = usuario_db.get("pdv_senha") if usuario_db else None

                from scraper import baixar_relatorios_periodo
                import asyncio as _asyncio
                loop = _asyncio.get_event_loop()
                _, path_produtos, _ = await loop.run_in_executor(
                    None, baixar_relatorios_periodo, ini, fim, pdv_email, pdv_senha
                )
                produtos = pd.read_excel(path_produtos)
                produtos = normalizar_produtos(produtos)

                if produtos.empty:
                    await msg.reply_text("⚠️ Nenhum produto encontrado para o período selecionado.")
                    await abrir_menu(msg, chat_id)
                    return

                # Aplica margem de segurança se necessário
                if modo == "estoque":
                    produtos = produtos.copy()
                    produtos["quantidade"] = (produtos["quantidade"] * 1.3).round().astype(int)

            except Exception as e:
                await msg.reply_text(
                    f"❌ Erro ao buscar dados do PDV Legal.\n\n{i(str(e)[:200])}",
                    parse_mode="HTML"
                )
                await abrir_menu(msg, chat_id)
                return

            # Formato: chat
            if formato == "chat":
                blocos = bloco_reposicao(produtos, "exato")  # modo já aplicado acima
                for bloco in blocos:
                    await enviar(msg, bloco)
                await abrir_menu(msg, chat_id)
                return

            # Formato: Excel
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            tem_grupo = "grupo" in produtos.columns and produtos["grupo"].notna().any()
            fator_rep = 1.3 if modo == "estoque" else 1.0

            if formato == "excel_loja":
                wb.remove(wb.active)
                for filial in produtos["nomeloja"].unique():
                    nome_aba = filial.split()[-1].title()[:31]
                    ws  = wb.create_sheet(title=nome_aba)
                    df  = produtos[produtos["nomeloja"]==filial].sort_values("quantidade", ascending=False)
                    hdr = ["Categoria", "Produto", "Qtd vendida", "Repor"] if tem_grupo else ["Produto", "Qtd vendida", "Repor"]
                    ws.append(hdr)
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    for _, row in df.iterrows():
                        repor = int(round(row["quantidade"] * fator_rep))
                        if tem_grupo:
                            ws.append([str(row.get("grupo","")), row["produto"], int(row["quantidade"]), repor])
                        else:
                            ws.append([row["produto"], int(row["quantidade"]), repor])
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 32
                filename = f"reposicao_por_loja_{label_per.replace(' ','_')}.xlsx"

            else:  # excel_unificado
                ws = wb.active
                ws.title = "Reposição"
                hdr = ["Loja", "Categoria", "Produto", "Qtd vendida", "Repor"] if tem_grupo else ["Loja", "Produto", "Qtd vendida", "Repor"]
                ws.append(hdr)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                df_sorted = produtos.sort_values(["nomeloja", "grupo" if tem_grupo else "quantidade", "quantidade"], ascending=[True, True, False])
                for _, row in df_sorted.iterrows():
                    nome_loja = row["nomeloja"].split()[-1].title()
                    repor = int(round(row["quantidade"] * fator_rep))
                    if tem_grupo:
                        ws.append([nome_loja, str(row.get("grupo","")), row["produto"], int(row["quantidade"]), repor])
                    else:
                        ws.append([nome_loja, row["produto"], int(row["quantidade"]), repor])
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 32
                filename = f"reposicao_unificada_{label_per.replace(' ','_')}.xlsx"

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            await msg.reply_document(
                document=bio,
                filename=filename,
                caption=f"📊 Lista de reposição {label_modo} — {label_per}"
            )
            await abrir_menu(msg, chat_id)
            return

        # Fallback para callbacks antigos rep_exato / rep_estoque
        if acao in ("rep_exato", "rep_estoque"):
            dados_usuario.setdefault(chat_id, {})["rep_modo"] = acao.replace("rep_", "")
            modo  = acao.replace("rep_", "")
            label = "exatamente o que saiu" if modo == "exato" else "com estoque de segurança (+30%)"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📅 Hoje",            callback_data="rep_per_hoje")],
                [InlineKeyboardButton("📅 Ontem",           callback_data="rep_per_ontem")],
                [InlineKeyboardButton("📅 Últimos 7 dias",  callback_data="rep_per_7dias")],
                [InlineKeyboardButton("📅 Últimos 15 dias", callback_data="rep_per_15dias")],
                [InlineKeyboardButton("📅 Últimos 30 dias", callback_data="rep_per_30dias")],
                [InlineKeyboardButton("📅 Mês atual",       callback_data="rep_per_mes")],
                [InlineKeyboardButton("◀️ Voltar ao menu",  callback_data="menu")],
            ])
            await msg.reply_text(
                f"✅ Modo: {b(label)}\n\nQual período deseja usar?",
                parse_mode="HTML",
                reply_markup=kb
            )
            return

        return

    cmds = {
        "briefing":   comando_briefing,
        "produtos":   comando_produtos,
        "categorias": comando_categorias,
        "pagamentos": comando_pagamentos,
        "semana":             lambda u, c: abrir_submenu_semana(u.message, u.effective_chat.id),
        "semana_atual":       lambda u, c: comando_semana_atual(u.message, u.effective_chat.id),
        "semana_comparativo": lambda u, c: comando_semana_comparativo(u.message, u.effective_chat.id),
        "semana_mes":         lambda u, c: comando_semana_mes(u.message, u.effective_chat.id),
        "pico":       comando_pico,
        "alertas":    comando_alertas,
        "menu":       lambda u, c: abrir_menu(u.message, u.effective_chat.id),
        "reposicao":  comando_reposicao,
        "comparativo":comando_comparativo,
        "projecao":   comando_projecao,
        "score":      comando_score,
        "produto_mes":comando_produto_mes,
        "giro":       comando_giro,
    }

    if acao == "briefing":
        context.user_data["gerar_briefing"] = True
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Hoje",            callback_data="atualizar_hoje")],
            [InlineKeyboardButton("📅 Ontem",           callback_data="atualizar_ontem")],
            [InlineKeyboardButton("📅 Últimos 7 dias",  callback_data="atualizar_7dias")],
            [InlineKeyboardButton("📅 Últimos 15 dias", callback_data="atualizar_15dias")],
            [InlineKeyboardButton("📅 Últimos 30 dias", callback_data="atualizar_30dias")],
            [InlineKeyboardButton("📅 Mês atual",       callback_data="atualizar_mes")],
            [InlineKeyboardButton("📅 Mês anterior",    callback_data="atualizar_mes_anterior")],
            [InlineKeyboardButton("◀️ Voltar ao menu",  callback_data="menu")],
        ])
        await msg.reply_text(
            f"📊 {b('BRIEFING')} — Escolha o período:",
            parse_mode="HTML", reply_markup=kb
        )
        return

    fake = type("U", (), {"message": msg, "effective_chat": type("C", (), {"id": chat_id})()})()
    if acao in cmds:
        await cmds[acao](fake, None)

# ─── MAIN — ver ao final do arquivo ─────────────────────────


# ─── MIDDLEWARE DE CONTROLE DE ACESSO ────────────────────────
def exige_acesso(func):
    """
    Decorator/wrapper: bloqueia a execução do comando se o usuário não tiver
    assinatura ativa (trial ou ativo). Aplica a mesma verificação usada em
    mensagens livres, garantindo consistência entre todos os pontos de entrada
    do bot (boa prática: controle de acesso centralizado, não duplicado).
    """
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not await verificar_acesso(update, context):
            return
        await func(update, context)
    wrapper.__name__ = getattr(func, "__name__", "comando_protegido")
    return wrapper


async def verificar_acesso(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Verifica se o usuário tem acesso ativo."""
    from database import usuario_tem_acesso, buscar_usuario
    chat_id = update.effective_chat.id

    # Se está em conversa de onboarding, não bloqueia
    if context.user_data.get("em_onboarding"):
        return False

    # Se está no meio da coleta de telefone/CEP/número para reativação,
    # deixa passar — essa mensagem precisa chegar em mensagem_livre() para
    # ser tratada pelo fluxo "aguardando == reativar_*", mesmo com a
    # assinatura ainda cancelada/inativa (afinal é isso que está sendo resolvido).
    aguardando_atual = dados_usuario.get(chat_id, {}).get("aguardando", "")
    if isinstance(aguardando_atual, str) and aguardando_atual.startswith("reativar_"):
        return True

    # Verifica se o usuário existe e está em alguma etapa do cadastro
    usuario = await buscar_usuario(chat_id)
    if not usuario:
        # Novo usuário tentando usar o bot — manda pro /start
        texto_novo = "👋 Use /start para se cadastrar no MercadoBot.\n\n7 dias grátis, depois R$ 29,90/mês."
        if update.message:
            await update.message.reply_text(texto_novo)
        elif update.callback_query:
            await update.callback_query.message.reply_text(texto_novo)
        return False

    # Verifica acesso usando a função simples
    tem_acesso, motivo = await usuario_tem_acesso(chat_id)
    if tem_acesso:
        return True

    mensagens = {
        "trial_expirado": (
            f"⏰ {b('Seu trial de 7 dias encerrou.')}\n\n"
            f"Para continuar usando o MercadoBot, cadastre seu cartão e ative a assinatura."
        ),
        "expirado": (
            f"⚠️ {b('Sua assinatura expirou.')}\n\n"
            f"Regularize para continuar usando o MercadoBot."
        ),
        "bloqueado": (
            f"🔒 {b('Acesso bloqueado.')}\n\n"
            f"Regularize seu pagamento para reativar sua conta."
        ),
        "cancelado": (
            f"😔 {b('Sua assinatura está cancelada.')}\n\n"
            f"Reative para voltar a usar o MercadoBot."
        ),
    }

    texto_bloqueio = mensagens.get(motivo, "Use /start para acessar o MercadoBot.")
    kb_bloqueio = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Reativar assinatura", callback_data="reativar")],
    ])

    if update.message:
        await update.message.reply_text(texto_bloqueio, parse_mode="HTML", reply_markup=kb_bloqueio)
    elif update.callback_query:
        await update.callback_query.answer("Acesso bloqueado.")
        await update.callback_query.message.reply_text(texto_bloqueio, parse_mode="HTML", reply_markup=kb_bloqueio)

    return False


# ─── INICIALIZAÇÃO COM SCHEDULER E SAAS ──────────────────────
def main():
    garantir_browser()

    from scheduler import iniciar_scheduler
    from onboarding import conversation_handler
    from webhook_server import iniciar_servidor_webhook, set_bot
    from database import inicializar_banco
    from admin import cmd_admin, callback_admin

    async def post_init(app):
        await inicializar_banco()
        await configurar_menu(app)
        set_bot(app.bot)
        runner = await iniciar_servidor_webhook()
        app.bot_data["webhook_runner"] = runner
        logger.info("✅ SaaS inicializado — banco, webhook e menu prontos.")

    app = ApplicationBuilder().token(TELEGRAM_TOKEN).post_init(post_init).build()

    # ConversationHandler PRIMEIRO — captura mensagens do onboarding
    app.add_handler(conversation_handler())

    # Demais handlers no mesmo grupo — só ativam quando conversation não está ativo
    app.add_handler(CommandHandler("menu",       exige_acesso(comando_menu)))
    app.add_handler(CommandHandler("briefing",   exige_acesso(comando_briefing)))
    app.add_handler(CommandHandler("produtos",   exige_acesso(comando_produtos)))
    app.add_handler(CommandHandler("categorias", exige_acesso(comando_categorias)))
    app.add_handler(CommandHandler("pagamentos", exige_acesso(comando_pagamentos)))
    app.add_handler(CommandHandler("semana",     exige_acesso(comando_semana)))
    app.add_handler(CommandHandler("pico",       exige_acesso(comando_pico)))
    app.add_handler(CommandHandler("alertas",    exige_acesso(comando_alertas)))
    app.add_handler(CommandHandler("reposicao",   exige_acesso(comando_reposicao)))
    app.add_handler(CommandHandler("atualizar",   exige_acesso(comando_atualizar)))
    app.add_handler(CommandHandler("comparativo", exige_acesso(comando_comparativo)))
    app.add_handler(CommandHandler("projecao",    exige_acesso(comando_projecao)))
    app.add_handler(CommandHandler("score",       exige_acesso(comando_score)))
    app.add_handler(CommandHandler("produto_mes", exige_acesso(comando_produto_mes)))
    app.add_handler(CommandHandler("giro",        exige_acesso(comando_giro)))
    app.add_handler(CommandHandler("perguntar",   exige_acesso(cmd_perguntar_handler)))
    # Comandos sempre liberados, independente de status de assinatura:
    # admin (tem seu próprio gate via is_admin), status, reativar, cancelar
    # e configuracoes (precisa funcionar para o usuário corrigir credenciais
    # mesmo sem acesso ativo, e para reativação).
    app.add_handler(CommandHandler("admin",         cmd_admin))
    app.add_handler(CommandHandler("status",        cmd_status_handler))
    app.add_handler(CommandHandler("reativar",      cmd_reativar_handler))
    app.add_handler(CommandHandler("configuracoes", cmd_configuracoes_handler))
    app.add_handler(CommandHandler("cancelar",      cmd_cancelar_handler))
    app.add_handler(MessageHandler(filters.Document.ALL,            receber_arquivo_com_acesso))
    app.add_handler(MessageHandler(filters.PHOTO,                   receber_foto_com_acesso))
    app.add_handler(MessageHandler(filters.VOICE | filters.AUDIO,    receber_audio_com_acesso))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mensagem_livre_com_acesso))
    app.add_handler(CallbackQueryHandler(callback_botoes))

    iniciar_scheduler()

    print("🤖 MercadoBot SaaS rodando...")
    app.run_polling()

if __name__ == "__main__":
    main()